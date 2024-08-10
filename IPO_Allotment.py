# C:\Python311\python.exe -m pip install wmi     #using this command to install module in python 'this command is use when python is more than 1'

import customtkinter as ctk
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from tkinter import filedialog
from datetime import date
from wmi import WMI
import wmi
import sys
import tkinter as tk
from tkinter import ttk
import pandas as pd
import openpyxl
import time
import keyboard
import mysql.connector
from PIL import Image
import os
import requests
import subprocess
import json
import base64
import zipfile
from datetime import datetime
import io
from PIL import Image,ImageTk
from tkinter import messagebox
import ver
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives import serialization, hashes
import traceback
import hashlib
import getmac


def toggle_theme():
    current_theme = ctk.get_appearance_mode().strip()  # Remove leading/trailing whitespace
    if current_theme == 'Light':
        ctk.set_appearance_mode("dark")
    else:
        ctk.set_appearance_mode("light")

ctk.set_appearance_mode("light")
ctk.set_default_color_theme('gray')

root = ctk.CTk()
root.title("IPO Allotment")
root.geometry("1020x520")

icon_path = 'trade.png'
original_icon = Image.open(icon_path)
resized_icon = original_icon.resize((16, 16), Image.BICUBIC)

tk_icon = ImageTk.PhotoImage(resized_icon)

root.iconphoto(True, tk_icon)   


#kfintech Links
kfintechlink1 = "https://kosmic.kfintech.com/ipostatus/"
kfintechlink2 = "https://kcas.kfintech.com/ipostatus/"
kfintechlink3 = "https://kprism.kfintech.com/ipostatus/"
kfintechlink4 = "https://evault.kfintech.com/ipostatus/"
kfintechlink5 = "https://rti.kfintech.com/ipostatus/"

#BigShare Links
bigsharelink1 = "https://ipo.bigshareonline.com/IPO_Status.html"
bigsharelink2 = "https://ipo1.bigshareonline.com/IPO_Status.html"
bigsharelink3 = "https://ipo2.bigshareonline.com/IPO_Status.html"

# from selenium.webdriver.common.keys import Keys
current_date = date.today()

def check_internet_connection():
    try:
        requests.get("http://www.google.com", timeout=5)
        return True
    except requests.ConnectionError:
        return False
    
if not check_internet_connection():
        messagebox.showerror("Connection Error", "E-01 \n Please check your internet connection.")
        sys.exit()
        # return

#Check for Update in Application #WITH PROGRESS BAR
def check_update():
    try:
        # GETTING LATEST VERSION FROM GITHUB
        url = 'https://api.github.com/repos/{owner}/{repo}/contents/{path}'
        owner = 'devamtsanghvi'
        repo = 'IPO-Allotment'
        path = 'version.txt'
        params = {'ref': 'main'}

        response = requests.get(url.format(owner=owner, repo=repo, path=path), params=params)
        content = json.loads(response.content)
        # print(content)

        file_content = base64.b64decode(content['content']).decode('utf-8')
        print(file_content)

        if ver.version == file_content:
            print("Latest Version")
        else:
            response = messagebox.askyesno('Update Available', 'An update is available. Do you want to download and install it?')
            if response == tk.YES:
                try:
                    import os
                    import shutil

                    def remove_files_except_some(folder_path, files_to_preserve, folders_to_preserve):
                        # Ensure the folder exists
                        if not os.path.exists(folder_path):
                            return
                        
                        # Get a list of all files and directories in the folder
                        all_files = os.listdir(folder_path)
                        
                        # Iterate through all files and directories in the folder
                        for item_name in all_files:
                            item_path = os.path.join(folder_path, item_name)
                            
                            # Check if the item is a file
                            if os.path.isfile(item_path):
                                # Check if the file is not in the preserve list
                                if item_name not in files_to_preserve:
                                    try:
                                        os.remove(item_path)
                                    except Exception as e:
                                        pass
                            
                            # Check if the item is a directory
                            elif os.path.isdir(item_path):
                                # Check if the directory is not in the preserve list
                                if item_name not in folders_to_preserve:
                                    try:
                                        shutil.rmtree(item_path)
                                    except Exception as e:
                                            pass
                    # Example usage
                    folder_path =  r'C:\IPO_Allotment'
                    files_to_preserve = ['update.exe','IPO_Allotment.exe','base_library.zip']
                    folders_to_preserve = ['certifi']  

                    remove_files_except_some(folder_path, files_to_preserve, folders_to_preserve)
                except :
                    pass
                    # print(e)

                url1 = 'https://api.github.com/repos/{owner}/{repo}/releases/latest'
                # Send the GET request to retrieve the release information
                response = requests.get(url1.format(owner=owner, repo=repo))
                release_info = json.loads(response.content)

                new_dialog = tk.Toplevel()
                new_dialog.geometry("300x100")
                new_dialog.title("Update Downloading")

                # Add a progress bar to the new dialog box
                progress_bar = ttk.Progressbar(new_dialog, orient="horizontal", length=200, mode="determinate")
                progress_bar.pack(pady=20)

                # Add a label to show the progress percentage
                progress_label = tk.Label(new_dialog, text="Downloading update (0%)")
                progress_label.pack()

                # Download the release assets
                r = requests.get(release_info['zipball_url'], stream=True)
                zipfile_bytes = io.BytesIO()
                total_size = int(r.headers.get('Content-Length', sys.maxsize)) # Set an arbitrary large value if Content-Length is not provided
                # chunk_size = 2048 # 2 kB
                chunk_size = 4194304 # 4 MB
                downloaded_size = 0
                if(os.path.exists("update.zip")):
                    os.remove("update.zip")
                with open("update.zip", "wb") as f:
                    for chunk in r.iter_content(chunk_size=chunk_size):
                        if chunk:
                            zipfile_bytes.write(chunk)
                            print("Downloaded size: ", downloaded_size)
                            print("CHUNK SIZE: ", len(chunk))
                            print("TOTAL SIZE: ", total_size)
                            downloaded_size += len(chunk)
                            progress_pct = int(downloaded_size / total_size * 100)
                            progress_bar['value'] = progress_pct
                            progress_label.config(text=f"Downloading update ({progress_pct}%)")
                            print(f"Downloading update ({progress_pct}%")
                            new_dialog.update()

                #Remove ZIP file
                os.remove("update.zip")
                new_dialog.destroy()

                if(os.path.exists("update.exe")):
                    os.remove("update.exe")

                # Extract the exe file from the release assets
                with zipfile.ZipFile(zipfile_bytes) as z:
                    for filename in z.namelist():
                        if filename.endswith('.exe'):
                            with open("update.exe", 'wb') as f:
                                f.write(z.read(filename))
                            break

                with open('run_update.bat', 'w') as f:
                    f.write('start update.exe')

                subprocess.call('run_update.bat', shell=True)

                os.remove("run_update.bat")
                sys.exit(0)
    except:
        messagebox.showerror("Connection Error", f"E-02 \n Connection Error")
        sys.exit()

check_update()

def load_private_key(file_path):
    with open(file_path, "rb") as f:
        private_key = serialization.load_pem_private_key(
            f.read(),
            password=None,
            backend=default_backend()
        )
    return private_key

def fetch_encrypted_content(url, params):
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()  # Raise exception for bad responses
        content = json.loads(response.content)
        encrypted_content = base64.b64decode(content['content'])
        print(encrypted_content)
        return encrypted_content
    except (requests.RequestException, json.JSONDecodeError) as e:
        print("Error fetching encrypted content:", e)
        return None

def decrypt_file(private_key, encrypted_content):
    try:
        decrypted_text = private_key.decrypt(
            encrypted_content,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
        return decrypted_text
    except Exception as e:
        print(e)

def read_database_config1(private_key):
    encrypted_content = fetch_encrypted_content(url, params)
    if encrypted_content is None:
        return None

    lines = decrypt_file(private_key, encrypted_content).decode().split('\n')
    
    return {
        'host': lines[0].strip(),
        'user': lines[1].strip(),
        'password': lines[2].strip(),
        'database': lines[3].strip(),
    }

def read_database_config2(private_key):
    encrypted_content = fetch_encrypted_content(url, params)
    if encrypted_content is None:
        return None

    lines = decrypt_file(private_key, encrypted_content).decode().split('\n')
    
    return {
        'host': lines[4].strip(),
        'user': lines[5].strip(),
        'password': lines[6].strip(),
        'database': lines[7].strip(),
    }

private_key = load_private_key('private_key.pem')
url = 'https://api.github.com/repos/Dipak19-b/Ipo_Allotment/contents/next.txt.enc'
params = {'ref': 'main'}

def dbconnection():
    def server2():  
        for _ in range(3):
            try:
                config = read_database_config2(private_key)     
                mydb = mysql.connector.connect(**config)
                return mydb
            except Exception as e:
                time.sleep(0.5)  
        return None

    def server1():  
        for _ in range(3):
            try:
                config = read_database_config1(private_key)
                mydb = mysql.connector.connect(**config)
                return mydb
            except Exception as e:
                print(e)
                time.sleep(0.5)  
        return None
    
    mydb = server1()
    if mydb is None:
        mydb = server2()
        if mydb is None:
            messagebox.showerror("Connection Error", "E-03 \n Connection Error.")
            sys.exit()

    return mydb

mydb = dbconnection() 

def fetch_data():
    global data
    conn = mydb
    cursor = conn.cursor()

    macid = WMI().Win32_ComputerSystemProduct()[0].UUID
    
    GetMacid = f"select MacId from Customer where MacId = '{macid}'"
    df = pd.read_sql(GetMacid, con=conn)

    try:
        result = df.iloc[0, 0]
    except:
        return None,None
    
    if result == macid:
        cursor.execute(f"select ExpiryDate from Customer where MacId = '{macid}'")
        data = cursor.fetchall()
        data = data[0]
        data = data[0]
        today = datetime.now().date()
        days_remaining = (data - today).days
        days_remaining = f"{days_remaining} Days left"
        
        GetMacid = f"select Name from Customer where MacId = '{macid}'"
        df = pd.read_sql(GetMacid, con=conn)

        user_name = df.iloc[0, 0]
        print(user_name)

        return days_remaining,user_name
    else:
        return None,None 

data,user_name  = fetch_data()

def captcha_solution(browser,CaptchaSolutionAPI,image_path):
    try:
        l = browser.find_element("xpath",
                image_path)
        browser.execute_script("arguments[0].scrollIntoView();", l)
        l = l.screenshot_as_png
        
    except Exception as e: 
        print(e)    
        l = browser.find_element("xpath",
                image_path)
        browser.execute_script("arguments[0].scrollIntoView();", l)
        l = l.screenshot_as_png
                 
    img = Image.open(io.BytesIO(l))
    img = img.resize((200, 50))
    img = img.crop((40, 0, 160, 50))
    img = img.resize((200, 50))
    
    resized_image_bytes = io.BytesIO()
    img.save(resized_image_bytes, format='PNG')
    base64_resized = base64.b64encode(resized_image_bytes.getvalue()).decode('utf-8')

    dataa = json.dumps({"image_base": base64_resized,"type":"kfintech"}) #kfintech

    try:
        response = requests.get(CaptchaSolutionAPI,data =dataa )
        response = response.json() 
        stqw = response['body']
        return stqw
    except:
        pass

def BigShare_captcha_solution(browser,CaptchaSolutionAPI):
    l = browser.find_element("xpath",
            '/html/body/div/div/div[3]/div/div/div/div[1]/div/div/div/div[8]/canvas').screenshot_as_png
    
    image = Image.open(io.BytesIO(l))
    resized_image = image.resize((200, 50))
    resized_image_bytes = io.BytesIO()
    resized_image.save(resized_image_bytes, format='PNG')
    base64_resized = base64.b64encode(resized_image_bytes.getvalue()).decode('utf-8')

    dataa = json.dumps({"image_base": base64_resized,'type':'bigshare'}) #bigshare

    try:
        response = requests.get(CaptchaSolutionAPI,data =dataa )
        response = response.json()
        stqw = response['body']
        return stqw
    except:
        pass

#Function To Run Karv[Old]
def kary(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url):
    ErrorText = "ERROR"
    df = pd.read_excel(path)
    # browser = webdriver.Chrome()
    browser.implicitly_wait(5)
    browser.maximize_window()
    
    browser.get(url)
    browser.execute_script("document.body.style.zoom='100%'")
    
    browser.execute_script("document.documentElement.style.height = '335px';")
    browser.execute_script("document.documentElement.style.width = '1900px';")
    #Select IPO Name
    ipo_display = browser.find_element("xpath",'/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[2]/label/span').click()
    
    ipo = browser.find_element("xpath",
        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[3]/select')
    # ipo.click()
    ipo.send_keys(iponame)

    #Select Pan Number
    pan = browser.find_element("xpath",
        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[4]/div[3]/input')
    browser.execute_script("arguments[0].click();", pan)
    # pan.click()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    # mr = ws1.max_row
    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0
    continueVar = 0
    mycursor = mydb.cursor()
    #Today's Date
    value = date.today().strftime('%Y-%m-%d')

    #Get Retry Count
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1

    #Get Counter Date
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]

    #Check Counter Date
    if(Counterdate == None):
        #Update Counterdate If counterdate==none
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        #Update Counter 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    
    #If Counter Date is Equal To Today's Date
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0]) #Run as it is
    else:
        counter = 0
        #For Different Counter Date
        #Update Counter Date
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)

        #Update Counter
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    
    #Counter Total Counter
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    GetKaryAutomation = f"select KaryAutomation from Customer where Name = '{UserNameValue}'"
    dff = pd.read_sql(GetKaryAutomation, con=mydb)
    KaryAutomation = dff.iloc[0, 0]

    #Multiple Pan Value
    GetMultiplePan = f"select  MultiplePan from Customer where Name = '{UserNameValue}'"
    dff = pd.read_sql(GetMultiplePan, con=mydb)
    MultiplePan = dff.iloc[0, 0]
    
    GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
    df3 = pd.read_sql(GetCounterlimite, con=mydb)
    counterlimit = int(df3.iloc[0, 0])
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1    

    for i in range(StartRowEntryValue, mr + 1):
        cflag = 1
        bflag = 1
        counter = counter + 1
        Totalcounter = Totalcounter + 1
        # count = count + 1
        # variable.set(str(count))
        # root.update()
        try:
            AvgSharePerApp1 = (totalshare/(count-error)) #Average Share
            AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
        except:
            AvgSharePerApp = 0
        # AvgSharePerApp = ((counter-error)/alloted) * lotsize
        variable4.set(str(AvgSharePerApp))
        # root.update()

        root.attributes("-topmost", True)

        #Check Counter Per Day
        if(counter >= counterlimit):
            #Update Counter
            AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddCounter)

            #Update Total Counter
            AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddTotalCounter)
            mydb.commit()
            
            tk.messagebox.showinfo(
                "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
            sys.exit()
        try:
            # print(ws1.cell(row = i, column = ei).value)
            panno = ws1.cell(row=i, column=PanEntryValue).value
            try:
                panno = panno.strip()
            except:
                panno = panno
            # print(panno)
            
            #Pan No Select
            pan = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[7]/div/input')
            pan.clear()
            pan.send_keys(panno)

            # time.sleep(1)
            if KaryAutomation == 1 or KaryAutomation == '1':
                GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                CaptchaSolutionAPI = dff.iloc[0, 0]
                for j in range(1, 15):

                    #Automatic Captcha
                    # time.sleep(2)
                    cap = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[8]/div[1]/input')
                    # time.sleep(2)
                    image_path = '//*[@id="captchaimg"]'
                    t = captcha_solution(browser,CaptchaSolutionAPI,image_path)
                    cap.clear()
                    cap.send_keys(t)
                    # print('3')
                    # time.sleep(1)
                    submit = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[9]/a')
                    # submit.click()
                    browser.execute_script("arguments[0].click();", submit)
                    browser.implicitly_wait(0)
                    try:
                        time.sleep(0.5)
                        novaildpan = browser.find_element("xpath",
                            '/html/body/div/div[2]/div/div/div/div/div/div/div/div[3]/div/div')
                        ErrorText = novaildpan.text
                        
                        if str(ErrorText) != str('Captcha is invalid.'): #Check Error Text of Invalid Captcha
                            if str(ErrorText) != str('CAPTCHA is invalid or Expired'): #Check Error Text of Expire or Invalid Captcha
                                if str(ErrorText) != str('Please enter Captcha.'): #Check Error Text of Enter Captcha
                                    RetryCount = 1
                                    continueVar = 1
                                    break
                                else:
                                    #Refresh Captcha
                                    browser.find_element("xpath",
                                        '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                                    refresh = browser.find_element("xpath",
                                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                            else:
                                #Refresh Captcha
                                browser.find_element("xpath",
                                    '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                                refresh = browser.find_element("xpath",
                                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                        else:
                            #Refresh Captcha
                            browser.find_element("xpath",
                                '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                            refresh = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                    except:
                        ErrorText = ''
                        break
                    browser.implicitly_wait(5)
            else:
                #For manually Add Captcha-[KaryAutomation-0] Condition
                cap = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[8]/div[1]/input')
                cap.click()
                keyboard.wait('tab')
                submit = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[9]/a')
                # submit.click()
                browser.execute_script("arguments[0].click();", submit)
                
                try:
                    novaildpan = browser.find_element("xpath",
                        '/html/body/div/div[2]/div/div/div/div/div/div')
                    ErrorText = novaildpan.text
                except:
                    ErrorText = "ERROR"
                    pass

            if continueVar == 1:
                browser.find_element("xpath","/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button").click()
                error = error + 1
                count = count + 1
                variable.set(str(count))
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = novaildpan.text
                df.to_excel(path, index=False)
                continueVar = 0
                bflag = 0
                continue
            #Get Value
            value = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[7]/strong/span').text
            value1 = int(value)
            Name = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[3]/strong/span').text
            
            #Get DP ID
            DPIDClientID1 = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[4]/strong/span').text

            #Get Application Number
            Applicationnumber1 = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[1]/strong/span').text

            #Get Category
            Catetory1 = browser.find_element("xpath",
                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[2]/strong/span').text

            browser.implicitly_wait(0)

            #Multiple Pan
            if MultiplePan == 1 or MultiplePan == '1':
                try:
                    #Get Value
                    valuee = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[7]/strong/span').text
                    value2 = int(valuee)

                    #Get Name
                    Name2 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[3]/strong/span').text

                    #Get Dp Id
                    DPIDClientID2 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[4]/strong/span').text

                    #Get Application Number
                    Applicationnumber2 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[1]/strong/span').text

                    #Get Category
                    Catetory2 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[2]/strong/span').text
                    df.at[i-2, 'QTY of MultiplePan'] = value2
                    df.at[i-2, 'Name of MultiplePan '] = Name2
                    df.at[i-2, 'DP ID Client ID of MultiplePan'] = DPIDClientID2
                    df.at[i-2, 'Application Number of MultiplePan'] = Applicationnumber2
                    df.at[i-2, 'Category of MultiplePan'] = Catetory2

                except:
                    value2 = ''
                    Name2 = ''
                    DPIDClientID2 = ''
                    Applicationnumber2 = ''
                    Catetory2 = ''
            browser.implicitly_wait(5)
            if value1 > 0:
                cflag = 0
                alloted = alloted + 1
                variable1.set(str(alloted))
                totalshare = totalshare + int(value1)
                variable5.set(str(totalshare))

                root.update()

            if value1 == 0:
                cflag = 0
                not_alloted = not_alloted + 1
                variable2.set(str(not_alloted))
                root.update()
                
            count = count + 1
            variable.set(str(count))

            try:
                AvgSharePerApp1 = (totalshare/(count-error)) #Average Share
                AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
            except:
                AvgSharePerApp = 0
            variable4.set(str(AvgSharePerApp))
            
            root.update()

            #Add Value In Excel
            df.at[i-2, 'QTY'] = value1
            df.at[i-2, 'Name'] = Name
            df.at[i-2, 'DP ID Client ID'] = DPIDClientID1
            df.at[i-2, 'Application number'] = Applicationnumber1
            df.at[i-2, 'Catetory'] = Catetory1
            #If Get 2nd Box- GET Value

            df.to_excel(path, index=False)
            
            browser.implicitly_wait(0)
            try:
                another_query = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[2]/a')
                another_query.click()
            except:
                browser.execute_script(
                    "window.scrollTo(0, window.scrollY + 200)")
                another_query = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[2]/a')
                another_query.click()
            browser.implicitly_wait(5)
            
        except:

            if RetryCount == 1:
                if bflag == 1:
                    error =  error + 1
                    count = count + 1
                    variable.set(str(count))
                    variable3.set(str(error))
                
                    root.update()

                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = ErrorText
                df.to_excel(path, index=False)

                browser.quit()

                # browser = webdriver.Chrome()
                browser = get_webdriver(driver_name)
                browser.maximize_window()

                browser.get(url)
                
                browser.execute_script("document.body.style.zoom='100%'")
                # time.sleep(2)
                
                browser.execute_script("document.documentElement.style.height = '335px';")
                browser.execute_script("document.documentElement.style.width = '1900px';")
                # time.sleep(2)
                
                ipo_display = browser.find_element("xpath",'/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[2]/label/span').click()
    
                ipo = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[3]/select')
                ipo.send_keys(iponame)

                pan = browser.find_element("xpath",
                    '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[4]/div[3]/input')
                browser.execute_script("arguments[0].click();", pan)
            for j in range(1, RetryCount):
                try:
                    browser.quit()

                    # browser = webdriver.Chrome()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()

                    browser.get(url)
                    browser.execute_script("document.body.style.zoom='100%'")
                    # time.sleep(2)
                    
                    browser.execute_script("document.documentElement.style.height = '335px';")
                    browser.execute_script("document.documentElement.style.width = '1900px';")
                    
                    # time.sleep(2)
                    
                    ipo_display = browser.find_element("xpath",'/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[2]/label/span').click()
    
                    ipo = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[3]/select')
                    # ipo.click()
                    ipo.send_keys(iponame)

                    pan = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[4]/div[3]/input')
                    browser.execute_script("arguments[0].click();", pan)

                    panno1 = ws1.cell(row=i, column=PanEntryValue).value

                    pan = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[7]/div/input')
                    pan.clear()
                    pan.send_keys(panno)
                    # time.sleep(10)
                    # root.attributes("-topmost", False)
                    if KaryAutomation == 1 or KaryAutomation == '1':
                        GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                        dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                        CaptchaSolutionAPI = dff.iloc[0, 0]
                        for j in range(1, 6):
                            cap = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[8]/div[1]/input')
                            image_path = '//*[@id="captchaimg"]'
                            t = captcha_solution(browser,CaptchaSolutionAPI,image_path)
                            cap.clear()
                            cap.send_keys(t)
                            print('4')

                            submit = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[9]/a')
                            submit.click()
                            
                            browser.implicitly_wait(0)
                            try:
                                time.sleep(0.5)
                                novaildpan = browser.find_element("xpath",
                                    '/html/body/div/div[2]/div/div/div/div/div/div/div/div[3]/div/div')
                                ErrorText = novaildpan.text
                                if str(ErrorText) != str('Captcha is invalid.'):
                                    if str(ErrorText) != str('CAPTCHA is invalid or Expired'):
                                        if str(ErrorText) != str('Please enter Captcha.'):
                                            continueVar = 1
                                            break
                                        else:
                                            print('5')
                                            browser.find_element("xpath",
                                                '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                                            refresh = browser.find_element("xpath",
                                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                                    else:
                                        print('5')
                                        browser.find_element("xpath",
                                            '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                                        refresh = browser.find_element("xpath",
                                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                                else:
                                    print('6')
                                    browser.find_element("xpath",
                                        '/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button').click()
                                    refresh = browser.find_element("xpath",
                                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[8]/div[2]/a/i').click()
                            except:
                                ErrorText = ''
                                break
                            browser.implicitly_wait(5)
                    else:
                        cap = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div[1]/div/div/div/div[8]/div[1]/input')
                        cap.click()
                        keyboard.wait('tab')
                        submit = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[9]/a')
                        submit.click()
                        try:
                            novaildpan = browser.find_element("xpath",
                                '/html/body/div/div[2]/div/div/div/div/div/div')
                            ErrorText = novaildpan.text

                        except:
                            ErrorText = "ERROR"
                            pass

                    if continueVar == 1:
                        browser.find_element("xpath","/html/body/div/div[2]/div/div/div/div/div/div/div/div[4]/button").click()
                        eerror = error + 1
                        variable3.set(str(error))
                        count = count + 1
                        variable.set(str(count))
                        root.update()
                        df.at[i-2, 'QTY'] = novaildpan.text
                        df.to_excel(path, index=False)
                            
                        continueVar = 0
                        continue

                    value = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[7]/strong/span').text
                    value1 = int(value)
                    Name = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[3]/strong/span').text
                    DPIDClientID1 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[4]/strong/span').text

                    Applicationnumber1 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[1]/strong/span').text

                    Catetory1 = browser.find_element("xpath",
                        '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td/div/div/div/div[2]/div[2]/strong/span').text
                    browser.implicitly_wait(0)
                    if MultiplePan == 1 or MultiplePan == '1':
                        try:
                            valuee = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[7]/strong/span').text
                            value2 = int(valuee)
                            Name2 = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[3]/strong/span').text

                            DPIDClientID2 = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[4]/strong/span').text

                            Applicationnumber2 = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[1]/strong/span').text

                            Catetory2 = browser.find_element("xpath",
                                '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[3]/td/div/div/div/div[2]/div[2]/strong/span').text
                            df.at[i-2, 'QTY of MultiplePan'] = value2
                            df.at[i-2, 'Name of MultiplePan '] = Name2
                            df.at[i-2, 'DP ID Client ID of MultiplePan'] = DPIDClientID2
                            df.at[i-2, 'Application Number of MultiplePan'] = Applicationnumber2
                            df.at[i-2, 'Category of MultiplePan'] = Catetory2
                            
                        except:
                            value2 = ''
                            Name2 = ''
                            DPIDClientID2 = ''
                            Applicationnumber2 = ''
                            Catetory2 = ''
                        browser.implicitly_wait(5)
                    if value1 > 0:
                        cflag = 0
                        alloted = alloted + 1
                        variable1.set(str(alloted))
                        totalshare = totalshare + int(value1)
                        variable5.set(str(totalshare))
                        root.update()

                    if value1 == 0:
                        cflag = 0
                        not_alloted = not_alloted + 1
                        variable2.set(str(not_alloted))
                        root.update()
                        
                    count = count + 1
                    variable.set(str(count))
                    
                    try:
                        AvgSharePerApp1 = (totalshare/(count-error)) 
                        AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
                    except:
                        AvgSharePerApp = 0
                    variable4.set(str(AvgSharePerApp))
                    
                    root.update()

                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Name'] = Name
                    df.at[i-2, 'DP ID Client ID'] = DPIDClientID1
                    df.at[i-2, 'Application number'] = Applicationnumber1
                    df.at[i-2, 'Catetory'] = Catetory1

                    df.to_excel(path, index=False)
                    browser.implicitly_wait(0)
                    try:
                        another_query = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[2]/a')
                        another_query.click()
                    except:
                        browser.execute_script(
                            "window.scrollTo(0, window.scrollY + 200)")
                        another_query = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div[2]/div/div[2]/a')
                        another_query.click()
                    browser.implicitly_wait(5)

                except:
                    if j >= RetryCount-1:
                        browser.quit()

                        # browser = webdriver.Chrome()
                        browser = get_webdriver(driver_name)
                        browser.maximize_window()

                        browser.get(url)
                        browser.execute_script("document.body.style.zoom='100%'")
                        # time.sleep(2)
                        
                        browser.execute_script("document.documentElement.style.height = '335px';")
                        browser.execute_script("document.documentElement.style.width = '1900px';")
                        
                        # time.sleep(2)
                        
                        ipo_display = browser.find_element("xpath",'/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[2]/label/span').click()
    
                        ipo = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[3]/select')
                        # ipo.click()
                        ipo.send_keys(iponame)

                        pan = browser.find_element("xpath",
                            '/html/body/form/div[3]/div[2]/div/div[2]/div/div/div/div/div/div[4]/div[3]/input')
                        browser.execute_script("arguments[0].click();", pan)
                        if cflag == 0:
                            count = count + 1
                            variable.set(str(count))
                            error = error + 1
                            variable3.set(str(error))
                            root.update()
                            
                        df.at[i-2, 'QTY'] = ErrorText
                        df.to_excel(path, index=False)
                        
        # variable.set(str(i-1))
    
    #Update Counter
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)

    #Update Total Counter
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
        
    browser.quit()

#Function For Linkin
def linkin(browser,driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url):
    ErrorText = "ERROR"
    df = pd.read_excel(path, engine='openpyxl')

    # browser = webdriver.Chrome()
    browser.implicitly_wait(5)
    browser.maximize_window()
    browser.get(url)

    #IPO Name
    ipo = browser.find_element("xpath", '//*[@id="ddlCompany"]')
    ipo.click()
    ipo.send_keys(iponame)

    loginbtn = browser.find_element("xpath", '//*[@id="349"]')
    loginbtn.click()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0
    continueVar = 0
    mycursor = mydb.cursor()

    #Today's Date
    value = date.today().strftime('%Y-%m-%d')

    #Select Retry Count
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1

    #Select Counter Date
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]

    #Check Counter Date
    if(Counterdate == None):

        #Set Counter Date if its Empty
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0

        #Update Counter
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    
    #Check Counter Date and Today's Date
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        #Update Counter Date
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)

        #Update Counter Start to 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    
    #Select Total Counter
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    #Select Multiple Pan
    GetMultiplePan = f"select  MultiplePan from Customer where Name = '{UserNameValue}'"
    dff = pd.read_sql(GetMultiplePan, con=mydb)
    MultiplePan = dff.iloc[0, 0]
    
    GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
    df3 = pd.read_sql(GetCounterlimite, con=mydb)
    counterlimit = int(df3.iloc[0, 0])
    
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1

    for i in range(StartRowEntryValue, mr + 1):
        count = count + 1
        Totalcounter = Totalcounter + 1
        variable.set(str(count))
        # browser.delete_all_cookies()
        try:
            AvgSharePerApp1 = (totalshare/(count-error))
            AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
        except:
            AvgSharePerApp = 0
        variable4.set(str(AvgSharePerApp))

        root.attributes("-topmost", True)
        counter = counter + 1
        
        if(counter >= counterlimit):
            AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddCounter)
            AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddTotalCounter)
            mydb.commit()
            tk.messagebox.showinfo(
                "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
            sys.exit()
        try:

            panno = (ws1.cell(row=i, column=PanEntryValue).value)
            try:
                panno = panno.strip()
            except:
                panno = panno

            #Select Pan
            pan = browser.find_element("xpath", '//*[@id="txtStat"]')
            pan.clear()
            pan.send_keys(panno)
            browser.implicitly_wait(0)
            try:
                Alert = WebDriverWait(browser, 1).until(EC.visibility_of_element_located(
                (By.XPATH, '/html/body/div[6]/div[2]/div/p/label')))
                ErrorText = Alert.text
                RetryCount = 1
                continueVar = 1
            except:
                ErrorText = "ERROR"
            browser.implicitly_wait(5)

            SUBMIT = browser.find_element(By.ID, 'btnsearc')
            SUBMIT.click()

            browser.implicitly_wait(0)

            try:
                Alert = WebDriverWait(browser, 1).until(EC.visibility_of_element_located(
                (By.XPATH, '/html/body/div[6]/div[2]/div/p/label')))
                            # /html/body/div[6]/div[2]/div/p/label
                ErrorText = Alert.text
                RetryCount = 1
                continueVar = 1
            except:
                ErrorText = "ERROR"
            browser.implicitly_wait(5)

            if continueVar == 1:
                try:
                    continueVar = 0
                    browser.find_element("xpath","/html/body/div[6]/div[3]/div/button").click()
                    error = error + 1
                    variable3.set(str(error))
                    root.update()
                    df.at[i-2, 'QTY'] = ErrorText
                    df.to_excel(path, index=False)
                    
                    continue
                    
                except:
                    continueVar = 0

            lastTable = len(browser.find_elements("xpath",
                "/html/body/div[3]/div/div/div[2]/div[3]/table"))
            
            value = browser.find_element("xpath",
                '/html/body/div[3]/div/div/div[2]/div[3]/table[1]/tbody/tr[4]/td[2]').text
            # print(value)
            value1 = value

            Name = browser.find_element("xpath",
                '/html/body/div[3]/div/div/div[2]/div[3]/table[1]/tbody/tr[2]/td[2]').text
            CutOffPrice = browser.find_element("xpath",
                '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[3]/td[4]').text
            SecuritiesApplied = browser.find_element("xpath",
                '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[3]/td[2]').text
            Category = browser.find_element("xpath",
                '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[1]/th/span').text
            Category = Category.split("-")[1]
            browser.implicitly_wait(0)
            try:
                err = browser.find_element("xpath",
                    '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[5]/td').text
                df.at[i-2, 'ERROR'] = err
            except:
                pass
            browser.implicitly_wait(5)

            try:
                if int(value1) > 0:
                    alloted = alloted + 1
                    variable1.set(str(alloted))
                    totalshare = totalshare + int(value1)
                    variable5.set(str(totalshare))

                    root.update()

                if int(value1) == 0:
                    not_alloted = not_alloted + 1
                    variable2.set(str(not_alloted))
                    root.update()
            except:
                error = error + 1
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = 'Error'
                df.to_excel(path, index=False)
                    
            browser.implicitly_wait(0)
            try:
                ErrorText = browser.find_element("xpath", '//*[@id="tbl_DetSec"]/table/tbody/tr[5]/td').text
                df.at[i-2, 'QTY'] = ErrorText
                df.to_excel(path, index=False)
                
            except:
                pass
            browser.implicitly_wait(5)
            df.at[i-2, 'QTY'] = value1 #Alloted Or Not Allotted
            df.at[i-2, 'Name'] = Name
            if Name == '':
                df.at[i-2, 'QTY'] = "ClientId/DpId Mismatch"
            df.at[i-2, 'Cut Off Price'] = CutOffPrice
            df.at[i-2, 'Securities Applied'] = SecuritiesApplied
            df.at[i-2, 'Category'] = Category

            for m in range(0, lastTable-1):
                if MultiplePan == 1 or MultiplePan == '1':

                    MultiplePanvalue = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[4]/td[2]').text
                    MultiplePanvalue1 = int(MultiplePanvalue)

                    MultiplePanName = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[2]/td[2]').text
                    MultiplePanCutOffPrice = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[3]/td[4]').text
                    MultiplePanSecuritiesApplied = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[3]/td[2]').text
                    MultipleCategory = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[1]/th/span').text
                    MultipleCategory = MultipleCategory.split("-")[1]
                    df.at[i-2, f'QTY of MultiplePan {m + 1}'] = MultiplePanvalue1
                    df.at[i-2, f'Name of MultiplePan {m + 1}'] = MultiplePanName
                    df.at[i-2, f'Cut Off Price of MultiplePan {m + 1}'] = MultiplePanCutOffPrice
                    df.at[i-2, f'Securities Applied of MultiplePan {m + 1}'] = MultiplePanSecuritiesApplied
                    df.at[i-2, f'Category of MultiplePan {m + 1}'] = MultipleCategory
                    
            df.to_excel(path, index=False)

        except:
            if RetryCount == 1:
                error = error + 1

                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = ErrorText
                df.to_excel(path, index=False)
                
                browser.quit()

                browser = get_webdriver(driver_name)
                browser.implicitly_wait(5)
                browser.maximize_window()
                browser.get(url)

                ipo = browser.find_element("xpath",
                    '//*[@id="ddlCompany"]')
                ipo.click()
                ipo.send_keys(iponame)

                loginbtn = browser.find_element("xpath", '//*[@id="349"]')
                loginbtn.click()
            for j in range(1, RetryCount):
                try:
                    browser.quit()

                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get(url)

                    ipo = browser.find_element("xpath",
                        '//*[@id="ddlCompany"]')
                    ipo.click()
                    ipo.send_keys(iponame)

                    loginbtn = browser.find_element("xpath",
                        '//*[@id="349"]')
                    loginbtn.click()

                    pan = browser.find_element("xpath", '//*[@id="txtStat"]')
                    pan.clear()
                    pan.send_keys(panno)
                    browser.implicitly_wait(0)

                    try:
                        Alert = WebDriverWait(browser, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[6]/div[2]/div/p/label')))
                        ErrorText = Alert.text
                        continueVar = 1
                    except:
                        ErrorText = "ERROR"
                    browser.implicitly_wait(5)

                    SUBMIT = browser.find_element(By.ID, 'btnsearc')
                    SUBMIT.click()

                    time.sleep(1)
                    browser.implicitly_wait(0)

                    try:
                        Alert = WebDriverWait(browser, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[6]/div[2]/div/p/label')))
                        ErrorText = Alert.text
                        continueVar = 1
                    except:
                        ErrorText = "ERROR"
                    browser.implicitly_wait(5)

                    if continueVar == 1:
                        browser.find_element("xpath","/html/body/div[6]/div[3]/div/button").click()
                        error = error + 1
                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = Alert.text
                        df.to_excel(path, index=False)
                    
                        continueVar = 0
                        continue

                    lastTable = len(browser.find_elements("xpath",
                        "/html/body/div[3]/div/div/div[2]/div[3]/table"))

                    value = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[1]/tbody/tr[4]/td[2]').text
                    Name = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table[1]/tbody/tr[2]/td[2]').text
                    CutOffPrice = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[3]/td[4]').text
                    SecuritiesApplied = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[3]/td[2]').text
                    Category = browser.find_element("xpath",
                        '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[1]/th/span').text
                    Category = Category.split("-")[1]
                    value1 = value
                    browser.implicitly_wait(0)
                    try:
                        err = browser.find_element("xpath",
                            '/html/body/div[3]/div/div/div[2]/div[3]/table/tbody/tr[5]/td').text
                        df.at[i-2, 'ERROR'] = err
                        
                    except:
                        pass
                    browser.implicitly_wait(5)

                    try:
                        if int(value1) > 0:
                            alloted = alloted + 1
                            variable1.set(str(alloted))
                            totalshare = totalshare + int(value1)
                            variable5.set(str(totalshare))
                            root.update()

                        if int(value1) == 0:
                            not_alloted = not_alloted + 1
                            variable2.set(str(not_alloted))
                            root.update()
                    except:
                        error = error + 1
                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = 'Error'
                        df.to_excel(path, index=False)
                        
                    browser.implicitly_wait(0)
                    try:
                        ErrorText = browser.find_element("xpath", '//*[@id="tbl_DetSec"]/table/tbody/tr[5]/td').text
                        df.at[i-2, 'QTY'] = ErrorText
                        df.to_excel(path, index=False)
                        
                    except:
                        pass
                    browser.implicitly_wait(5)

                    df.at[i-2, 'QTY'] = value1
                    if Name == '':
                        df.at[i-2, 'QTY'] = "ClientId/DpId Mismatch"
                    df.at[i-2, 'Name'] = Name
                    df.at[i-2, 'Cut Off Price'] = CutOffPrice
                    df.at[i-2, 'Securities Applied'] = SecuritiesApplied
                    df.at[i-2, 'Category'] = Category

                    for m in range(0, lastTable-1):
                        if MultiplePan == 1 or MultiplePan == '1':

                            MultiplePanvalue = browser.find_element("xpath",
                                '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[4]/td[2]').text
                            MultiplePanvalue1 = int(MultiplePanvalue)

                            MultiplePanName = browser.find_element("xpath",
                                '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[2]/td[2]').text
                            MultiplePanCutOffPrice = browser.find_element("xpath",
                                '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[3]/td[4]').text
                            MultiplePanSecuritiesApplied = browser.find_element("xpath",
                                '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[3]/td[2]').text
                            MultipleCategory = browser.find_element("xpath",
                                '/html/body/div[3]/div/div/div[2]/div[3]/table[2]/tbody/tr[1]/th/span').text
                            MultipleCategory = MultipleCategory.split("-")[1]
                            df.at[i-2, f'QTY of MultiplePan {m + 1}'] = MultiplePanvalue1
                            df.at[i-2, f'Name of MultiplePan {m + 1}'] = MultiplePanName
                            df.at[i-2, f'Cut Off Price of MultiplePan {m + 1}'] = MultiplePanCutOffPrice
                            df.at[i-2, f'Securities Applied of MultiplePan {m + 1}'] = MultiplePanSecuritiesApplied
                            df.at[i-2, f'Category of MultiplePan {m + 1}'] = MultipleCategory

                    df.to_excel(path, index=False)

                except:
                    if j >= RetryCount-1:
                        browser.quit()

                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        browser.get(url)

                        ipo = browser.find_element("xpath",
                            '//*[@id="ddlCompany"]')
                        ipo.click()
                        ipo.send_keys(iponame)

                        loginbtn = browser.find_element("xpath",
                            '//*[@id="349"]')
                        loginbtn.click()

                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = ErrorText
                        df.to_excel(path, index=False)

    
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
    
    browser.quit()

#FUnction For Bigshare
def bigshare(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url):
    t = 0
    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0
    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetKaryAutomation = f"select KaryAutomation from Customer where Name = '{UserNameValue}'"
    dff = pd.read_sql(GetKaryAutomation, con=mydb)
    KaryAutomation = dff.iloc[0, 0]
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])
    GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
    df3 = pd.read_sql(GetCounterlimite, con=mydb)
    counterlimit = int(df3.iloc[0, 0])

    for i in range(StartRowEntryValue, mr + 1):
        count = count + 1
        Totalcounter = Totalcounter + 1
        variable.set(str(count))
        try:
            AvgSharePerApp1 = (totalshare/(count-error))
            AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
        except:
            AvgSharePerApp = 0
        variable4.set(str(AvgSharePerApp))

        root.attributes("-topmost", True)
        counter = counter + 1
        
        if(counter >= counterlimit):
            AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddCounter)
            AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddTotalCounter)
            mydb.commit()
            tk.messagebox.showinfo(
                "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
            sys.exit()

        try:
            browser.get(url)
            try :
                browser.implicitly_wait(0)                
                ads = WebDriverWait(browser, 0).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[1]/div/span'))
                    )
                ads.click()
            except:
                pass
            browser.implicitly_wait(5)  
            
            WebDriverWait(browser, 30).until(EC.visibility_of_element_located(
                (By.XPATH, '/html/body/div/div/div[3]/div/div/div/div[1]/div/div/div/div[1]/label')))
            ipo = browser.find_element("xpath",
                '//*[@id="ddlCompany"]')
            ipo.send_keys(iponame)

            panselect = browser.find_element("xpath",
                '//*[@id="ddlSelectionType"]')
            panselect.send_keys("PAN Number")

            panno = (ws1.cell(row=i, column=PanEntryValue).value)

            pan = browser.find_element("xpath",
                '//*[@id="txtpan"]')
            pan.clear()
            pan.send_keys(panno)

            if KaryAutomation == 1 or KaryAutomation == '1':
                GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                CaptchaSolutionAPI = dff.iloc[0, 0]
                for j in range(1, 15):
                    #Automatic Captcha
                    cap = browser.find_element("xpath",'//*[@id="captcha-input"]')
                    cap.click()
                    # t = BigShare_captcha_solution(browser,CaptchaSolutionAPI)
                    cap.clear()
                    captcha_code = browser.execute_script("return sessionStorage.getItem('captchaCode');")
                    cap.send_keys(captcha_code)
                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))
                    browser.implicitly_wait(0)
                    # time.sleep(1)
                    novaildpan = browser.find_element("xpath",
                        '//*[@id="lblcaptcha"]')
                    ErrorText = novaildpan.text
                    if ErrorText != "":
                        
                        #Refresh Captcha
                        refresh = browser.find_element("xpath",
                            '//*[@id="refresh-captcha"]').click()
            
                    else:
                        break
                    browser.implicitly_wait(5)
            else:
                #For manually Add Captcha-[KaryAutomation-0] Condition
                cap = browser.find_element("xpath",'//*[@id="captcha-input"]')
                cap.click()
                keyboard.wait('tab')
                browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))
                
                novaildpan = browser.find_element("xpath",
                    '//*[@id="lblcaptcha"]')
                ErrorText = novaildpan.text
                if ErrorText != "":
                    ErrorText = novaildpan.text
                else:
                    ErrorText = "ERROR"
                    pass
            try:
                e = browser.find_element("xpath",'/html/body/div[4]/h2')
                sta = e.text
                ok = browser.find_element("xpath",'/html/body/div[4]/div[7]/div/button').click()
                error = error + 1
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = sta
                df.to_excel(path, index=False)
                
                continue
            except:
                pass
            browser.implicitly_wait(5)
            Name = browser.find_element("xpath",
                '//*[@id="lbl3"]').text
            value1 = browser.find_element("xpath",
                '//*[@id="lbl5"]').text
            if value1 == 'NON-ALLOTTE':
                value1 = 0
            ClientIdDpId = browser.find_element("xpath",
                '//*[@id="lbl2"]').text
            AppliedShare = browser.find_element("xpath",
                '//*[@id="lbl4"]').text
            if int(value1) > 0:
                alloted = alloted + 1
                variable1.set(str(alloted))
                totalshare = totalshare + int(value1)
                variable5.set(str(totalshare))
                root.update()

            if int(value1) == 0:
                not_alloted = not_alloted + 1
                variable2.set(str(not_alloted))
                root.update()
            df.at[i-2, 'QTY'] = value1
            df.at[i-2, 'Name'] = Name
            df.at[i-2, 'ClientId/DpId'] = ClientIdDpId
            df.at[i-2, 'Applied Share'] = AppliedShare
            df.to_excel(path, index=False)
            
        except:
            if RetryCount == 1:
                error = error + 1

                variable3.set(str(error))
                root.update()
                if ErrorText == "":          
                    ErrorText = "Error"
                df.at[i-2, 'QTY'] = ErrorText
                df.to_excel(path, index=False)
                

                browser.quit()
                browser = get_webdriver(driver_name)
                browser.implicitly_wait(5)
                browser.maximize_window()

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get(url)
                    try :
                        browser.implicitly_wait(0)
                        ads = WebDriverWait(browser, 0).until(
                                EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div/div/div[1]/div/span'))
                            )
                        ads.click()
                    except:
                        pass
                    browser.implicitly_wait(5)
                    WebDriverWait(browser, 30).until(EC.visibility_of_element_located(
                        (By.XPATH, '/html/body/div/div/div[3]/div/div/div/div[1]/div/div/div/div[1]/label')))
                    ipo = browser.find_element("xpath",
                        '//*[@id="ddlCompany"]')
                    ipo.send_keys(iponame)

                    panselect = browser.find_element("xpath",
                        '//*[@id="ddlSelectionType"]')
                    panselect.send_keys("PAN Number")

                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    try:
                        panno = panno.strip()
                    except:
                        panno = panno
                    pan = browser.find_element("xpath",
                        '//*[@id="txtpan"]')
                    pan.clear()
                    pan.send_keys(panno)

                    if KaryAutomation == 1 or KaryAutomation == '1':
                        GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                        dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                        CaptchaSolutionAPI = dff.iloc[0, 0]
                        for j in range(1, 15):

                            #Automatic Captcha
                            cap = browser.find_element("xpath",'//*[@id="captcha-input"]')
                            cap.click()
                            # t = BigShare_captcha_solution(browser,CaptchaSolutionAPI)
                            cap.clear()
                            captcha_code = browser.execute_script("return sessionStorage.getItem('captchaCode');")
                            cap.send_keys(captcha_code)
                            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))
                            browser.implicitly_wait(0)
                            novaildpan = browser.find_element("xpath",
                                '//*[@id="lblcaptcha"]')
                            ErrorText = novaildpan.text
                            if ErrorText != "":
                                #Refresh Captcha
                                refresh = browser.find_element("xpath",
                                    '//*[@id="refresh-captcha"]').click()
                    
                            else:
                                break
                            browser.implicitly_wait(5)
                    else:
                        #For manually Add Captcha-[KaryAutomation-0] Condition
                        cap = browser.find_element("xpath",'//*[@id="captcha-input"]')
                        cap.click()
                        keyboard.wait('tab')
                        browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))
                        novaildpan = browser.find_element("xpath",
                            '//*[@id="lblcaptcha"]')
                        ErrorText = novaildpan.text
                        if ErrorText != "":
                            ErrorText = novaildpan.text
                        else:
                            ErrorText = "ERROR"
                            pass
                    try:
                        e = browser.find_element("xpath",'/html/body/div[4]/h2')
                        sta = e.text
                        ok = browser.find_element("xpath",'/html/body/div[4]/div[7]/div/button').click()
                        error = error + 1
                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = sta
                        df.to_excel(path, index=False)
                        
                        continue
                    except:
                        pass
                    browser.implicitly_wait(5)

                    WebDriverWait(browser, 3).until(EC.visibility_of_element_located(
                        (By.XPATH, '//*[@id="lbl5"]')))
                    Name = browser.find_element("xpath",
                        '//*[@id="lbl3"]').text

                    value1 = browser.find_element("xpath",
                        '//*[@id="lbl5"]').text
                    if value1 == 'NON-ALLOTTE':
                        value1 = 0
                    ClientIdDpId = browser.find_element("xpath",
                        '//*[@id="lbl2"]').text
                    AppliedShare = browser.find_element("xpath",
                        '//*[@id="lbl4"]').text
                    if int(value1) > 0:
                        alloted = alloted + 1
                        variable1.set(str(alloted))
                        totalshare = totalshare + int(value1)
                        variable5.set(str(totalshare))
                        root.update()

                    if int(value1) == 0:
                        not_alloted = not_alloted + 1
                        variable2.set(str(not_alloted))
                        root.update()
                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Name'] = Name
                    df.at[i-2, 'ClientId/DpId'] = ClientIdDpId
                    df.at[i-2, 'Applied Share'] = AppliedShare
                    df.to_excel(path, index=False)
                    
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = "Error"
                        df.to_excel(path, index=False)
    
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
                   
    browser.quit()

#Function For Sky Line
def skyline(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue):

    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare=0

    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    for i in range(StartRowEntryValue, mr + 1):
        count = count + 1
        Totalcounter = Totalcounter + 1
        variable.set(str(count))
        try:
            AvgSharePerApp1 = (totalshare/(count-error))
            AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
        except:
            AvgSharePerApp = 0
        variable4.set(str(AvgSharePerApp))

        root.attributes("-topmost", True)
        counter = counter + 1
        

        GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
        df3 = pd.read_sql(GetCounterlimite, con=mydb)
        counterlimit = int(df3.iloc[0, 0])
        if(counter >= counterlimit):
            AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddCounter)
            AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
            mycursor.execute(AddTotalCounter)
            mydb.commit()
            tk.messagebox.showinfo(
                "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
            sys.exit()

        try:
            browser.get('https://www.skylinerta.com/ipo.php')
            ipo = browser.find_element("xpath",
                '//*[@id="company"]')
            ipo.send_keys(iponame)

            panno = (ws1.cell(row=i, column=PanEntryValue).value)

            pan = browser.find_element("xpath",
                '//*[@id="pan"]')
            pan.clear()
            pan.send_keys(panno)

            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="display_application"]/ul/li[7]/span/input[3]'))))

            try:
                Name = browser.find_element("xpath",
                    '/html/body/div[2]/div/div[2]/div/p[2]')
                name1 = Name.text
                txt = name1.split(" : ", 1)
                text1 = txt[1]
                print(text1)
                value1 = browser.find_element("xpath",
                    '/html/body/div[2]/div/div[2]/div/div/table/tbody/tr[2]/td[3]').text
                print(value1)
                if value1 == 'NON-ALLOTTE':
                    value1 = 0
                ClientIdDpId = browser.find_element("xpath",
                    '/html/body/div[2]/div/div[2]/div/p[3]')
                cId = ClientIdDpId.text
                cdId = cId.split(" : ", 1)
                cdId1 = cdId[1]
                print(cdId1)
                AppliedShare = browser.find_element("xpath",
                    '/html/body/div[2]/div/div[2]/div/div/table/tbody/tr[2]/td[1]').text
            except:
                error = error + 1
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = "Error"
                df.to_excel(path, index=False)
                
                continue

            if int(value1) > 0:
                alloted = alloted + 1
                variable1.set(str(alloted))
                totalshare = totalshare + int(value1)
                variable5.set(str(totalshare))
                root.update()

            if int(value1) == 0:
                not_alloted = not_alloted + 1
                variable2.set(str(not_alloted))
                root.update()
            df.at[i-2, 'QTY'] = value1
            df.at[i-2, 'Name'] = text1
            df.at[i-2, 'ClientId/DpId'] = cdId1
            df.at[i-2, 'Applied Share'] = AppliedShare
            df.to_excel(path, index=False)
            

        except:
            if RetryCount == 1:
                error = error + 1

                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = "Error"
                df.to_excel(path, index=False)

                browser.quit()
                browser = get_webdriver(driver_name)
                browser.implicitly_wait(5)
                browser.maximize_window()

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get('https://www.skylinerta.com/ipo.php')
                    ipo = browser.find_element("xpath",
                        '//*[@id="company"]')
                    ipo.send_keys(iponame)

                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    try:
                        panno = panno.strip()
                    except:
                        panno = panno
                    pan = browser.find_element("xpath",
                        '//*[@id="pan"]')
                    pan.clear()
                    pan.send_keys(panno)

                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="display_application"]/ul/li[7]/span/input[3]'))))

                    try:
                        Name = browser.find_element("xpath",
                            '/html/body/div[2]/div/div[2]/div/p[2]')
                        name1 = Name.text
                        txt = name1.split(" : ", 1)
                        text1 = txt[1]
                        print(text1)

                        value1 = browser.find_element("xpath",
                            '/html/body/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td[5]').text
                        if value1 == 'NON-ALLOTTE':
                            value1 = 0
                        ClientIdDpId = browser.find_element("xpath",
                            '/html/body/div[2]/div/div[2]/div/p[3]')
                        cId = ClientIdDpId.text
                        cdId = cId.split(" : ", 1)
                        cdId1 = cdId[1]
                        print(cdId1)
                        AppliedShare = browser.find_element("xpath",
                            '/html/body/div/div/div[3]/div/div/div/div[3]/table/tbody/tr[2]/td[4]').text
                    except:
                        error = error + 1
                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = "Error"
                        df.to_excel(path, index=False)
                        
                        continue
                    if int(value1) > 0:
                        alloted = alloted + 1
                        variable1.set(str(alloted))
                        totalshare = totalshare + int(value1)
                        variable5.set(str(totalshare))
                        root.update()

                    if int(value1) == 0:
                        not_alloted = not_alloted + 1
                        variable2.set(str(not_alloted))
                        root.update()
                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Name'] = text1
                    df.at[i-2, 'ClientId/DpId'] = cdId
                    df.at[i-2, 'Applied Share'] = AppliedShare
                    df.to_excel(path, index=False)
                
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        df.at[i-2, 'QTY'] = "Error"
                        df.to_excel(path, index=False)
                        
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
        
    browser.quit()

def purva(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue):

    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0

    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    browser.get('https://www.purvashare.com/investor-service/ipo-query')

    for i in range(StartRowEntryValue, mr + 1):
        try:
            count = count + 1
            Totalcounter = Totalcounter + 1
            variable.set(str(count))

            try:
                AvgSharePerApp1 = (totalshare/(count-error))
                AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
            except:
                AvgSharePerApp = 0
            variable4.set(str(AvgSharePerApp))

            root.attributes("-topmost", True)
            counter = counter + 1

            GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
            df3 = pd.read_sql(GetCounterlimite, con=mydb)
            counterlimit = int(df3.iloc[0, 0])
            if(counter >= counterlimit):
                
                AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddCounter)
                AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddTotalCounter)
                mydb.commit()
                
                tk.messagebox.showinfo(
                    "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
                sys.exit()

            ipo = browser.find_element(By.XPATH,'//*[@id="company_id"]')
            ipo.send_keys(iponame)

            panno = (ws1.cell(row=i, column=PanEntryValue).value)

            pan = browser.find_element(By.XPATH,'//*[@id="ipo-query"]/p[4]/input')
            pan.clear()
            pan.send_keys(panno)

            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="ipo-query"]/p[5]/input'))))
            # time.sleep(2)
            
            try:
                Name = browser.find_element(By.XPATH,'/html/body/div[2]/main/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[1]')
                name1 = Name.text
            except:
                error_text = browser.find_element(By.XPATH,'//*[@id="content"]/div[1]/div/div/b').text
                
                df.at[i-2, 'QTY'] = error_text
                df.to_excel(path, index=False)
                
                error = error + 1

                variable3.set(str(error))
                root.update()
                continue

            App_num = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[2]').text
            
            value1 = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[6]').text
            if value1 == 'NON-ALLOTTE':
                value1 = 0
            ClientIdDpId = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[4]')
            cId = ClientIdDpId.text
            
            AppliedShare = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[5]').text

            if int(value1) > 0:
                alloted = alloted + 1
                variable1.set(str(alloted))
                totalshare = totalshare + int(value1)
                variable5.set(str(totalshare))
                root.update()

            if int(value1) == 0:
                not_alloted = not_alloted + 1
                variable2.set(str(not_alloted))
                root.update()
            
            df.at[i-2, 'QTY'] = value1
            df.at[i-2, 'Name'] = name1
            df.at[i-2, 'ClientId/DpId'] = cId
            df.at[i-2, 'Applied Share'] = AppliedShare
            df.at[i-2, 'Application Number'] = App_num
            df.to_excel(path, index=False)

        except:

            if RetryCount == 1:
                error = error + 1

                variable3.set(str(error))
                root.update()
                
                df.at[i-2, 'QTY'] = "error"
                df.to_excel(path, index=False)
                
                browser.back()

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get('https://www.purvashare.com/investor-service/ipo-query')
                    ipo = browser.find_element(By.XPATH,'//*[@id="company_id"]')
                    ipo.send_keys(iponame)

                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    try:
                        panno = panno.strip()
                    except:
                        panno = panno
                    pan = browser.find_element(By.XPATH,'//*[@id="ipo-query"]/p[4]/input')
                    pan.clear()
                    pan.send_keys(panno)

                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="ipo-query"]/p[5]/input'))))

                    try:
                        Name = browser.find_element(By.XPATH,'/html/body/div[2]/main/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[1]')
                        name1 = Name.text
                    except:
                        error_text = browser.find_element(By.XPATH,'//*[@id="content"]/div[1]/div/div/b').text
                        
                        df.at[i-2, 'QTY'] = error_text
                        df.to_excel(path, index=False)
                        
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        continue
                    App_num = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[2]').text

                    value1 = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[6]').text
                    if value1 == 'NON-ALLOTTE':
                        value1 = 0
                    ClientIdDpId = browser.find_element(By.XPATH,'//*[@id="content"]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[4]')
                    cId = ClientIdDpId.text
                    AppliedShare = browser.find_element(By.XPATH,"/html/body/div[2]/main/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[5]").text
                    if int(value1) > 0:
                        alloted = alloted + 1
                        variable1.set(str(alloted))
                        totalshare = totalshare + int(value1)
                        variable5.set(str(totalshare))
                        root.update()

                    if int(value1) == 0:
                        not_alloted = not_alloted + 1
                        variable2.set(str(not_alloted))
                        root.update()
                    
                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Name'] = name1
                    df.at[i-2, 'ClientId/DpId'] = cId
                    df.at[i-2, 'Applied Share'] = AppliedShare
                    df.at[i-2, 'Application Number'] = App_num
                    df.to_excel(path, index=False)
                    
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        error_text = browser.find_element(By.XPATH,'//*[@id="content"]/div[1]/div/div/b').text
                        
                        df.at[i-2, 'QTY'] = error_text
                        df.to_excel(path, index=False)
                        
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
    
    browser.quit()

def MaaShitla(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue):

    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0

    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    browser.get('https://maashitla.com/allotment-status/public-issues')

    for i in range(StartRowEntryValue, mr + 1):
        try:
            browser.implicitly_wait(5)
            
            count = count + 1
            Totalcounter = Totalcounter + 1
            variable.set(str(count))

            try:
                AvgSharePerApp1 = (totalshare/(count-error))
                AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
            except:
                AvgSharePerApp = 0
            variable4.set(str(AvgSharePerApp))

            root.attributes("-topmost", True)
            counter = counter + 1

            GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
            df3 = pd.read_sql(GetCounterlimite, con=mydb)
            counterlimit = int(df3.iloc[0, 0])
            if(counter >= counterlimit):
                
                AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddCounter)
                AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddTotalCounter)
                mydb.commit()
                
                tk.messagebox.showinfo(
                    "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
                sys.exit()

            ipo = browser.find_element(By.XPATH,'//*[@id="txtCompany"]')
            ipo.send_keys(iponame)

            panno = (ws1.cell(row=i, column=PanEntryValue).value)
            pan_check = browser.find_element(By.XPATH,'//*[@id="pan"]').click
            pan = browser.find_element(By.XPATH,'//*[@id="txtSearch"]')
            pan.clear()
            pan.send_keys(panno)

            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))
            # time.sleep(2)
            
            data = {}
            try:
                time.sleep(1)
                Name = browser.find_element('xpath','/html/body/div[1]/div/div/div/div/div/div[7]').text
                lines = Name.split('\n')
                for line in lines:
                    if ':' in line:
                        key, value = line.split(':', 1)
                        data[key.strip()] = value.strip()
                        
                browser.implicitly_wait(0)
                name1 = data['Name']
                App_num = data['Application No.']
                value1 = data['Share Alloted']
                AppliedShare = data['Share Applied']
                

                if int(value1) > 0:
                    alloted = alloted + 1
                    variable1.set(str(alloted))
                    totalshare = totalshare + int(value1)
                    variable5.set(str(totalshare))
                    root.update()

                if int(value1) == 0:
                    not_alloted = not_alloted + 1
                    variable2.set(str(not_alloted))
                    root.update()
                if data['Application No.'] == '':
                    df.at[i-2, 'QTY'] = 'Records Not Found...!!!'
                else:
                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Name'] = name1
                    df.at[i-2, 'Applied Share'] = AppliedShare
                    df.at[i-2, 'Application Number'] = App_num
                df.to_excel(path, index=False)
                
            except:
                error_text = browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div/div/div/div[8]').text
                
                df.at[i-2, 'QTY'] = error_text
                df.to_excel(path, index=False)
                
                error = error + 1

                variable3.set(str(error))
                root.update()
                continue
            

        except:
            if RetryCount == 1:
                error = error + 1

                variable3.set(str(error))
                root.update()
                
                df.at[i-2, 'QTY'] = "error"
                df.to_excel(path, index=False)
                
                browser.refresh()

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get('https://maashitla.com/allotment-status/public-issues')
                    
                    ipo = browser.find_element(By.XPATH,'//*[@id="txtCompany"]')
                    ipo.send_keys(iponame)
                    
                    pan_check = browser.find_element(By.XPATH,'//*[@id="pan"]').click
                    
                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    try:
                        panno = panno.strip()
                    except:
                        panno = panno
                    pan = browser.find_element(By.XPATH,'//*[@id="txtSearch"]')
                    pan.clear()
                    pan.send_keys(panno)

                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btnSearch"]'))))

                    data = {}
                    try:
                        time.sleep(1)
                        Name = browser.find_element('xpath','/html/body/div[1]/div/div/div/div/div/div[7]').text
                        lines = Name.split('\n')
                        for line in lines:
                            if ':' in line:
                                key, value = line.split(':', 1)
                                data[key.strip()] = value.strip()
                                
                        name1 = data['Name']
                        App_num = data['Application No.']
                        value1 = data['Share Alloted']
                        AppliedShare = data['Share Applied'] 
                        
                        if int(value1) > 0:
                            alloted = alloted + 1
                            variable1.set(str(alloted))
                            totalshare = totalshare + int(value1)
                            variable5.set(str(totalshare))
                            root.update()

                        if int(value1) == 0:
                            not_alloted = not_alloted + 1
                            variable2.set(str(not_alloted))
                            root.update()
                        
                        df.at[i-2, 'QTY'] = value1
                        df.at[i-2, 'Name'] = name1
                        df.at[i-2, 'Applied Share'] = AppliedShare
                        df.at[i-2, 'Application Number'] = App_num
                        df.to_excel(path, index=False)
                        
                    except:
                        error_text = browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div/div/div/div[8]').text
                        
                        df.at[i-2, 'QTY'] = error_text
                        df.to_excel(path, index=False)
                        
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        continue
                    
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        try:
                            error_text = browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div/div/div/div[8]').text
                            df.at[i-2, 'QTY'] = error_text
                        except:
                            df.at[i-2, 'QTY'] = 'Error'
                        
                        df.to_excel(path, index=False)
                        
                        browser.get('https://maashitla.com/allotment-status/public-issues')
                        
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
    
    browser.quit()

def Integrated(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue):

    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0

    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    browser.get('https://www.integratedregistry.in/RegistrarsToSTA.aspx?OD=2')

    for i in range(StartRowEntryValue, mr + 1):
        try:
            browser.implicitly_wait(5)
            
            count = count + 1
            Totalcounter = Totalcounter + 1
            variable.set(str(count))

            try:
                AvgSharePerApp1 = (totalshare/(count-error))
                AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
            except:
                AvgSharePerApp = 0
            variable4.set(str(AvgSharePerApp))

            root.attributes("-topmost", True)
            counter = counter + 1

            GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
            df3 = pd.read_sql(GetCounterlimite, con=mydb)
            counterlimit = int(df3.iloc[0, 0])
            if(counter >= counterlimit):
                
                AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddCounter)
                AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddTotalCounter)
                mydb.commit()
                
                tk.messagebox.showinfo(
                    "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
                sys.exit()

            ipo = browser.find_element(By.XPATH,'//*[@id="CompDdl2"]')
            ipo.send_keys(iponame)

            panno = (ws1.cell(row=i, column=PanEntryValue).value)
            pan_check = browser.find_element(By.XPATH,'//*[@id="ChoiceDdl2"]').send_keys('PAN Number')
            pan = browser.find_element(By.XPATH,'//*[@id="Pan_txt2"]')
            pan.clear()
            pan.send_keys(panno)

            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="NCD_sub_btn2"]'))))
            
            try:
                time.sleep(1)
                Name = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[6]/div[2]').text
                name1 = Name.strip(': ')
                APP_no = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[2]/div[2]').text
                App_num = APP_no.strip(': ')
                value = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[8]/div[2]').text
                value1 = value.strip(': ')
                AppliedShare = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[7]/div[2]').text
                AppliedShare = AppliedShare.strip(': ')
                Category = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[4]/div[2]').text
                Category = Category.strip(': ')
                DPid = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[5]/div[2]').text
                DPid = DPid.strip(': ')
                        
                browser.implicitly_wait(0)

                if int(value1) > 0:
                    alloted = alloted + 1
                    variable1.set(str(alloted))
                    totalshare = totalshare + int(value1)
                    variable5.set(str(totalshare))
                    root.update()

                if int(value1) == 0:
                    not_alloted = not_alloted + 1
                    variable2.set(str(not_alloted))
                    root.update()
                
                df.at[i-2, 'QTY'] = value1
                df.at[i-2, 'DPID'] = DPid
                df.at[i-2, 'Name'] = name1
                df.at[i-2, 'Applied Share'] = AppliedShare
                df.at[i-2, 'Category'] = Category
                df.at[i-2, 'Application Number'] = App_num
                df.to_excel(path, index=False)
                
                browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00"]/div[4]/section[6]/div[2]/div[3]/input'))))
            except :
                error_text = browser.find_element(By.XPATH,'/html/body/form/div[4]/section[6]/div[2]/div[2]/p').text
                
                df.at[i-2, 'QTY'] = error_text
                df.to_excel(path, index=False)
                
                error = error + 1

                variable3.set(str(error))
                root.update()
                
                browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00"]/div[4]/section[6]/div[2]/div[3]/input'))))
                
                continue
            

        except:
            if RetryCount == 1:
                error = error + 1
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = "error"
                df.to_excel(path, index=False)
                browser.back()

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get('https://www.integratedregistry.in/RegistrarsToSTA.aspx?OD=2')
                    
                    ipo = browser.find_element(By.XPATH,'//*[@id="CompDdl2"]')
                    ipo.send_keys(iponame)
                    
                    pan_check = browser.find_element(By.XPATH,'//*[@id="ChoiceDdl2"]').send_keys('PAN Number')
                    
                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    try:
                        panno = panno.strip()
                    except:
                        panno = panno
                    pan = browser.find_element(By.XPATH,'//*[@id="Pan_txt2"]')
                    pan.clear()
                    pan.send_keys(panno)

                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="NCD_sub_btn2"]'))))

                    data = {}
                    try:
                        time.sleep(1)
                        Name = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[6]/div[2]').text
                        name1 = Name.strip(': ')
                        APP_no = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[2]/div[2]').text
                        App_num = APP_no.strip(': ')
                        value = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[8]/div[2]').text
                        value1 = value.strip(': ')
                        AppliedShare = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[7]/div[2]').text
                        AppliedShare = AppliedShare.strip(': ')
                        Category = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[4]/div[2]').text
                        Category = Category.strip(': ')
                        DPid = browser.find_element('xpath','/html/body/form/div[4]/section[6]/div[2]/div[2]/div[5]/div[2]').text
                        DPid = DPid.strip(': ')
                                
                        browser.implicitly_wait(0)

                        if int(value1) > 0:
                            alloted = alloted + 1
                            variable1.set(str(alloted))
                            totalshare = totalshare + int(value1)
                            variable5.set(str(totalshare))
                            root.update()

                        if int(value1) == 0:
                            not_alloted = not_alloted + 1
                            variable2.set(str(not_alloted))
                            root.update()
                        
                        df.at[i-2, 'QTY'] = value1
                        df.at[i-2, 'DPID'] = DPid
                        df.at[i-2, 'Name'] = name1
                        df.at[i-2, 'Applied Share'] = AppliedShare
                        df.at[i-2, 'Category'] = Category
                        df.at[i-2, 'Application Number'] = App_num
                        df.to_excel(path, index=False)
                        
                        browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00"]/div[4]/section[6]/div[2]/div[3]/input'))))
                        
                    except:
                        error_text = browser.find_element(By.XPATH,'/html/body/form/div[4]/section[6]/div[2]/div[2]/p').text
                        
                        df.at[i-2, 'QTY'] = error_text
                        df.to_excel(path, index=False)
                        
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        
                        browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00"]/div[4]/section[6]/div[2]/div[3]/input'))))
                        
                        continue
                    
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        try:
                            error_text = browser.find_element(By.XPATH,'/html/body/form/div[4]/section[6]/div[2]/div[2]/p').text
                            df.at[i-2, 'QTY'] = error_text
                            
                            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="ctl00"]/div[4]/section[6]/div[2]/div[3]/input'))))
                        except:
                            df.at[i-2, 'QTY'] = 'Error'
                        
                        df.to_excel(path, index=False)
                        
                        browser.get('https://www.integratedregistry.in/RegistrarsToSTA.aspx?OD=2')
                        
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
    
    browser.quit()

def Cambridge(browser, driver_name, path, iponame, PanEntryValue, StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue):
    df = pd.read_excel(path)

    browser.implicitly_wait(5)
    browser.maximize_window()

    filename = path
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    if(StopRowEntryValue == "MAX ROW"):
        mr = ws1.max_row

    else:
        mr = int(StopRowEntryValue)
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    totalshare = 0

    GetKaryAutomation = f"select KaryAutomation from Customer where Name = '{UserNameValue}'"
    dff = pd.read_sql(GetKaryAutomation, con=mydb)
    KaryAutomation = dff.iloc[0, 0]

    mycursor = mydb.cursor()
    value = date.today().strftime('%Y-%m-%d')
    GetRetryCount = f"select RetryCount from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetRetryCount, con=mydb)
    RetryCount = int(df1.iloc[0, 0]) + 1
    GetCounterdate = f"select CounterDate from Customer where Name = '{UserNameValue}'"
    df1 = pd.read_sql(GetCounterdate, con=mydb)
    Counterdate = df1.iloc[0, 0]
    if(Counterdate == None):
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        counter = 0
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    if(str(Counterdate) == str(value)):
        GetCounter = f"select Counter from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetCounter, con=mydb)
        counter = int(df2.iloc[0, 0])
    else:
        counter = 0
        AddCounterDate = f"Update Customer set CounterDate = '{value}' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounterDate)
        AddCounter = f"Update Customer set Counter = '0' Where Name = '{UserNameValue}'"
        mycursor.execute(AddCounter)
        mydb.commit()
    GetTotalCounter = f"select TotalCounter from Customer where Name = '{UserNameValue}'"
    dff2 = pd.read_sql(GetTotalCounter, con=mydb)
    Totalcounter = int(dff2.iloc[0, 0])

    browser.get('https://ipo.cameoindia.com/')

    for i in range(StartRowEntryValue, mr + 1):
        try:
            browser.implicitly_wait(5)
            
            count = count + 1
            Totalcounter = Totalcounter + 1
            variable.set(str(count))

            try:
                AvgSharePerApp1 = (totalshare/(count-error))
                AvgSharePerApp = "{:.2f}".format(AvgSharePerApp1)
            except:
                AvgSharePerApp = 0
            variable4.set(str(AvgSharePerApp))

            root.attributes("-topmost", True)
            counter = counter + 1

            GetCounterlimite = f"select CounterLimit from Customer where Name = '{UserNameValue}'"
            df3 = pd.read_sql(GetCounterlimite, con=mydb)
            counterlimit = int(df3.iloc[0, 0])
            if(counter >= counterlimit):
                
                AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddCounter)
                AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddTotalCounter)
                mydb.commit()
                
                tk.messagebox.showinfo(
                    "Today's limit is over", "Today's limit is over \n Please contact administration on arhamtechnologyindia@gmail.com \n Whatsapp no:7228882088")
                sys.exit()

            ipo = browser.find_element("xpath",'//*[@id="drpCompany"]')
            ipo.send_keys(iponame)

            panno = (ws1.cell(row=i, column=PanEntryValue).value)
            pan_check = browser.find_element("xpath",'//*[@id="ddlUserTypes"]')
            pan_check.click()
            pan_check.send_keys("PAN NO")
            pan = browser.find_element("xpath",'//*[@id="txtfolio"]')
            pan.clear()
            pan.send_keys(panno)
            
            if KaryAutomation == 1 or KaryAutomation == '1':
                GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                CaptchaSolutionAPI = dff.iloc[0, 0]
                for j in range(1,15):
                    cap = browser.find_element("xpath",
                            '//*[@id="txt_phy_captcha"]')

                    image_path = '//*[@id="imgCaptcha"]'
                    
                    t = captcha_solution(browser,CaptchaSolutionAPI,image_path)
                    print(t)
                    cap.clear()
                    print(j)
                    cap.send_keys(t)
                    
                    submit = browser.find_element("xpath",'//*[@id="btngenerate"]')
                    browser.execute_script("arguments[0].click();", submit)
                    try:
                        time.sleep(2)
                        error_text = ''
                        while True:
                            try:
                                error_text = browser.find_element("xpath",'//*[@id="toast-container"]/div/div').text
                            except:
                                break
                        print(error_text)
                        error_msg = 'Oops!..Captcha entered is incorrect'
                        if str(error_text) == str(error_msg): 
                            time.sleep(1)
                            refresh = browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                EC.element_to_be_clickable(("xpath", '//*[@id="Button1"]'))))
                            time.sleep(1)
                            # ipo1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[2]/div/select')
                            ipo1 = browser.find_element("xpath",'//*[@id="drpCompany"]')
                            ipo1.click()
                            ipo1.send_keys(iponame)
                            time.sleep(1)
                            panno1 = (ws1.cell(row=i, column=PanEntryValue).value)
                            # pan_check1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[3]/div[1]/select')
                            pan_check1 = browser.find_element("xpath",'//*[@id="ddlUserTypes"]')
                            pan_check1.click()
                            pan_check1.send_keys("PAN NO")
                            time.sleep(1)
                            # pan1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[3]/div[2]/input')
                            pan1 = browser.find_element("xpath",'//*[@id="txtfolio"]')
                            
                            pan1.clear()
                            pan1.send_keys(panno1)
                            
                            continue
                        
                        elif str(error_text) == str('PAN NO should be 10 digit'):
                            df.at[i-2, 'QTY'] = error_text
                            df.to_excel(path, index=False)
                            
                            error = error + 1

                            variable3.set(str(error))
                            root.update()
                            refresh = browser.find_element("xpath",
                                '//*[@id="Button1"]').click()
                            # continue
                            break
                        elif str(error_text) == '':
                            break
                    except:
                        error_text = ''
                        break
                    
            try:
                time.sleep(1)
                name1 = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[1]').text
                print(name1)
                if name1 != str('NO DATA FOUND FOR THIS SEARCH KEY'):
                    value1 = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[2]').text
                    R_Amount = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[3]').text
                    R_mode = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[4]').text
                    Pj_No = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[5]').text
                    browser.implicitly_wait(0)

                    if int(value1) > 0:
                        alloted = alloted + 1
                        variable1.set(str(alloted))
                        totalshare = totalshare + int(value1)
                        variable5.set(str(totalshare))
                        root.update()

                    if int(value1) == 0:
                        not_alloted = not_alloted + 1
                        variable2.set(str(not_alloted))
                        root.update()
                    
                    df.at[i-2, 'Name'] = name1
                    df.at[i-2, 'QTY'] = value1
                    df.at[i-2, 'Refund_Amount'] = R_Amount
                    df.at[i-2, 'Refund_Mode'] = R_mode
                    df.at[i-2, 'Pj_No'] = Pj_No
                    df.to_excel(path, index=False)
                    
                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                    EC.element_to_be_clickable(("xpath", '//*[@id="Button1"]'))))
                else:
                    df.at[i-2, 'QTY'] = 'No Record Found'
                    df.to_excel(path, index=False)
                    
                    error = error + 1

                    variable3.set(str(error))
                    root.update()
                    
                    browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable(("xpath", '//*[@id="Button1"]'))))
                    
                    continue
                
            except :
                                                            # //*[@id="toast-container"]/div/div
                error_text = browser.find_element("xpath",'//*[@id="toast-container"]/div/div').text
                print(f'hlo-{error_text}')
                
                df.at[i-2, 'QTY'] = error_text
                df.to_excel(path, index=False)
                
                error = error + 1

                variable3.set(str(error))
                root.update()
                
                browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                    EC.element_to_be_clickable(("xpath", '//*[@id="Button1"]'))))
                
                continue
            

        except Exception as e:
            print(traceback.format_exc())
            if RetryCount == 1:
                error = error + 1
                variable3.set(str(error))
                root.update()
                df.at[i-2, 'QTY'] = "error"
                df.to_excel(path, index=False)
                browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="Button1"]'))))

            for j in range(1, RetryCount):
                try:
                    browser.quit()
                    browser = get_webdriver(driver_name)
                    browser.implicitly_wait(5)
                    browser.maximize_window()
                    browser.get('https://ipo.cameoindia.com/')
                    
                    ipo = browser.find_element("xpath",'//*[@id="drpCompany"]')
                    ipo.send_keys(iponame)

                    panno = (ws1.cell(row=i, column=PanEntryValue).value)
                    pan_check = browser.find_element("xpath",'//*[@id="ddlUserTypes"]')
                    pan_check.click()
                    pan_check.send_keys("PAN NO")
                    pan = browser.find_element("xpath",'//*[@id="txtfolio"]')
                    pan.clear()
                    pan.send_keys(panno)
                    
                    if KaryAutomation == 1 or KaryAutomation == '1':
                        GetCaptchaSolutionAPI = f"select CaptchaSolutionAPI from Customer where Name = '{UserNameValue}'"
                        dff = pd.read_sql(GetCaptchaSolutionAPI, con=mydb)
                        CaptchaSolutionAPI = dff.iloc[0, 0]
                        for j in range(1,15):
                            cap = browser.find_element("xpath",
                                    '//*[@id="txt_phy_captcha"]')

                            image_path = '//*[@id="imgCaptcha"]'
                            
                            t = captcha_solution(browser,CaptchaSolutionAPI,image_path)
                            print(t)
                            cap.clear()
                            print(j)
                            cap.send_keys('123456')
                            if j > 2:
                                cap.clear()
                                cap.send_keys(t)
                            
                            submit = browser.find_element("xpath",'//*[@id="btngenerate"]')
                            browser.execute_script("arguments[0].click();", submit)
                            try:
                                time.sleep(2)
                                error_text = ''
                                while True:
                                    try:
                                        error_text = browser.find_element("xpath",'//*[@id="toast-container"]/div/div').text
                                    except:
                                        break
                                print(error_text)
                                error_msg = 'Oops!..Captcha entered is incorrect'
                                if str(error_text) == str(error_msg): 
                                    time.sleep(1)
                                    refresh = browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                        EC.element_to_be_clickable(("xpath", '//*[@id="Button1"]'))))
                                    time.sleep(1)
                                    ipo1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[2]/div/select')
                                    ipo1.click()
                                    ipo1.send_keys(iponame)
                                    time.sleep(1)
                                    panno1 = (ws1.cell(row=i, column=PanEntryValue).value)
                                    pan_check1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[3]/div[1]/select')
                                    pan_check1.click()
                                    pan_check1.send_keys("PAN NO")
                                    time.sleep(1)
                                    pan1 = browser.find_element("xpath",'/html/body/form/div[3]/div/section/div[1]/div/div/div[3]/div[2]/input')
                                    pan1.clear()
                                    pan1.send_keys(panno1)
                                    
                                    continue
                                
                                elif str(error_text) == str('PAN NO should be 10 digit'):
                                    df.at[i-2, 'QTY'] = error_text
                                    df.to_excel(path, index=False)
                                    
                                    error = error + 1

                                    variable3.set(str(error))
                                    root.update()
                                    refresh = browser.find_element("xpath",
                                        '//*[@id="Button1"]').click()
                                    # continue
                                    break
                                elif str(error_text) == '':
                                    break
                            except:
                                error_text = ''
                                break
                    try:
                        time.sleep(1)
                        name1 = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[1]').text
                        if name1 != str('NO DATA FOUND FOR THIS SEARCH KEY'):
                            value1 = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[2]').text
                            R_Amount = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[3]').text
                            R_mode = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[4]').text
                            Pj_No = browser.find_element('xpath','/html/body/form/div[3]/div/section/div[3]/div/div/div/table/tbody/tr/td[5]').text
                            browser.implicitly_wait(0)

                            if int(value1) > 0:
                                alloted = alloted + 1
                                variable1.set(str(alloted))
                                totalshare = totalshare + int(value1)
                                variable5.set(str(totalshare))
                                root.update()

                            if int(value1) == 0:
                                not_alloted = not_alloted + 1
                                variable2.set(str(not_alloted))
                                root.update()
                            
                            df.at[i-2, 'Name'] = name1
                            df.at[i-2, 'QTY'] = value1
                            df.at[i-2, 'Refund_Amount'] = R_Amount
                            df.at[i-2, 'Refund_Mode'] = R_mode
                            df.at[i-2, 'Pj_No'] = Pj_No
                            df.to_excel(path, index=False)
                            
                            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="Button1"]'))))
                        
                        else:
                            df.at[i-2, 'QTY'] = 'No Record Found'
                            df.to_excel(path, index=False)
                            
                            error = error + 1

                            variable3.set(str(error))
                            root.update()
                            
                            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="Button1"]'))))
                            
                    except:
                        try:
                            error_text = browser.find_element(By.XPATH,'/html/body/form/div[4]/section[6]/div[2]/div[2]/p').text
                        except:
                            error_text = 'Error'
                        
                        df.at[i-2, 'QTY'] = error_text
                        df.to_excel(path, index=False)
                        
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        
                        browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="Button1"]'))))
                        
                        continue
                    
                except:
                    if j >= RetryCount-1:
                        browser.quit()
                        browser = get_webdriver(driver_name)
                        browser.implicitly_wait(5)
                        browser.maximize_window()
                        error = error + 1

                        variable3.set(str(error))
                        root.update()
                        try:
                            error_text = browser.find_element(By.XPATH,'/html/body/form/div[4]/section[6]/div[2]/div[2]/p').text
                            df.at[i-2, 'QTY'] = error_text
                            
                            browser.execute_script("arguments[0].click();", WebDriverWait(browser, 30).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="Button1"]'))))
                        except:
                            df.at[i-2, 'QTY'] = 'Error'
                        
                        df.to_excel(path, index=False)
                        
                        browser.get('https://ipo.cameoindia.com/')
                        
    AddCounter = f"Update Customer set Counter = '{counter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddCounter)
    AddTotalCounter = f"Update Customer set TotalCounter = '{Totalcounter}' Where Name = '{UserNameValue}'"
    mycursor.execute(AddTotalCounter)
    mydb.commit()
    
    browser.quit()

def get_webdriver(driver_name):
    try:
        if driver_name == "chrome":
            return webdriver.Chrome()
        elif driver_name == "firefox":
            return webdriver.Firefox()
        elif driver_name == "edge":
            return webdriver.Edge()
        else:
            raise tk.messagebox.showinfo("", "Please select proper browser from the Setting option")
    except:
        messagebox.showerror("E-08", 'There was an error while opening the Browser. Please try again.')

def submit():
    global variable,variable1,variable3,variable2,variable4,variable5
    alloted = 0
    not_alloted = 0
    error = 0
    count = 0
    AvgSharePerApp = 00.00
    totalshare = 0
    
    # mydb = dbconnection() 
    #username get From Tk Inter Value
    UserNameValue = UseNameEntry.get()
    mycursor = mydb.cursor()

    #Fetch Username From Database
    GetUserName = f"select * from Customer where Name = '{UserNameValue}'"
    mycursor.execute(GetUserName)   
    result = mycursor.fetchall()

    def main():
        #File Selection
        path = filedialog.askopenfilename(
            initialdir="/", title="select A files", filetype=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        try:   
            df = pd.read_excel(path)
            df.at[2-2, 'QTY'] = ''
            df.to_excel(path, index=False)
        except Exception as e:
            messagebox.showerror("E-07", 'Please close the Excel file and try again. The file is currently open and cannot be modified.')
            sys.exit()
        
        df = pd.read_excel(path)
        df.at[2-2, 'QTY'] = ''
        df.to_excel(path, index=False)
        # chromedriver_autoinstaller.install()

        #IPO Name From Tk Value
        IpoNameEntryValue = IpoNameEntry.get()

        #Get Pan Number From Excel File 
        PanEntryValue = int(PanEntry.get())

        #Start Row
        StartRowEntryValue = int(StartRowEntry.get())

        #Stop Row
        StopRowEntryValue = StopRowEntry.get()
        
        #Lot Size
        # lotsizevalue = int(lotsizeEntry.get())

        #Browser
        GetBrowser = f"select Browser from Customer where Name = '{UserNameValue}'"
        dfff2 = pd.read_sql(GetBrowser, con=mydb)
        driver_name = dfff2.iloc[0, 0]

        driver = get_webdriver(driver_name)

        counter()
        variable.set(str(count))
        variable1.set(str(alloted))
        variable2.set(str(not_alloted))
        variable3.set(str(error))
        variable4.set(str(AvgSharePerApp))
        variable5.set(str(totalshare))
        root.update()
        
        #Selection From The Radio Button
        if(var.get() == 1):

            #Get Karv URL
            GetKaryURL = f"select KaryURL from Customer where Name = '{UserNameValue}'"
            df2 = pd.read_sql(GetKaryURL, con=mydb)
            try:
                urll = int(df2.iloc[0, 0])
            except:
                url = kfintechlink1
            match urll:
                case 1:
                    url = kfintechlink1
                case 2:
                    url = kfintechlink2
                case 3:
                    url = kfintechlink3
                case 4:
                    url = kfintechlink4
                case 5:
                    url = kfintechlink5
                
            kary(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                    StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url)
        if(var.get() == 2):
            #Get Linkin URL
            GetLinkinURL = f"select LinkinURL from Customer where Name = '{UserNameValue}'"
            df2 = pd.read_sql(GetLinkinURL, con=mydb)
            url = df2.iloc[0, 0]
            print(url)
            linkin(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                   StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url)
        if(var.get() == 3):
            #Get Bigshare URL
            GetBigshareURL = f"select BigshareURL from Customer where Name = '{UserNameValue}'"
            df2 = pd.read_sql(GetBigshareURL, con=mydb)
            try:
                urll = int(df2.iloc[0, 0])
            except:
                url = bigsharelink1

            match urll:
                case 1:
                    url = bigsharelink1
                case 2:
                    url = bigsharelink2
                case 3:
                    url = bigsharelink3

            bigshare(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue, url)
        
        #Sky Line
        if(var.get() == 4):

            #Get Sky Line URL
            GetSkyLineURL = f"select BigshareURL from Customer where Name = '{UserNameValue}'"
            df2 = pd.read_sql(GetSkyLineURL, con=mydb)
            url = df2.iloc[0, 0]
            skyline(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue)
        #Purva
        if(var.get() == 5):
            purva(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue)
        
        if(var.get() == 6):
            MaaShitla(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue)
            
        if(var.get() == 7):
            Integrated(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue)
        if(var.get() == 8):
            Cambridge(driver, driver_name, path, IpoNameEntryValue, PanEntryValue,
                     StartRowEntryValue, StopRowEntryValue,  mydb, UserNameValue)

    if len(result): 
        #Expiry Date Get From Database
        GetExpiryDate = f"select ExpiryDate from Customer where Name = '{UserNameValue}'"
        df2 = pd.read_sql(GetExpiryDate, con=mydb)
        ExpiryDate = df2.iloc[0, 0]
        
        #Get Terms and Condition Value From Database
        GetTermsAndCondition = f"select TermsAndCondition from Customer where Name = '{UserNameValue}'"
        mycursor.execute(GetTermsAndCondition)
        TermsAndConditionvalue = mycursor.fetchall()
        strTermsAndConditionvalue = str(TermsAndConditionvalue[0][0])
        
        #Match Current Date With Expiry Date
        if(current_date >= ExpiryDate):

            tk.messagebox.showinfo(
                "E-05", "Your subscription has expired. Please contact administration at arhamtechnologyindia@gmail.com to renew your plan.\nWhatsApp: 7228882088")
            sys.exit()
        print("not empty")

        #Get Client MAC ID
        macid = WMI().Win32_ComputerSystemProduct()[0].UUID
        GetMacid = f"select MacId from Customer where Name = '{UserNameValue}'"
        df = pd.read_sql(GetMacid, con=mydb)
        result = df.iloc[0, 0]
        
        try:
            c = wmi.WMI()
            uuid = c.Win32_ComputerSystemProduct()[0].UUID
            mac_address = getmac.get_mac_address()
            serial_number = c.Win32_BIOS()[0].SerialNumber
            unique_string = f"{uuid}-{mac_address}-{serial_number}"
            unique_id = hashlib.sha256(unique_string.encode()).hexdigest()
            
            GetMacid = f"select Unique_Id from Customer where Name = '{UserNameValue}'"
            df = pd.read_sql(GetMacid, con=mydb)
            
            result1 = df.iloc[0, 0]
            
            if unique_id == result1:
                pass
            elif result1 == None:
                AddUnique_Id = f"Update Customer set Unique_Id = '{unique_id}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddUnique_Id)
                mydb.commit()
                main()
        except Exception as e:
            print(e)
            pass

        if macid == result:
            main()

        elif(result == None):
            TermsAndCondition = tk.messagebox.askquestion(
                "Terms and Conditions", strTermsAndConditionvalue, icon='question')
            print(TermsAndCondition)
            if TermsAndCondition == 'yes':
                AddMacid = f"Update Customer set MacId = '{macid}' Where Name = '{UserNameValue}'"
                mycursor.execute(AddMacid)
                mydb.commit()
                main()
            else:
                sys.exit()
        else:
            tk.messagebox.showinfo(
                "E-06", "This PC is not registered with your User Name. Please contact administration at arhamtechnologyindia@gmail.com to resolve this issue.\nWhatsApp: 7228882088")
            sys.exit()
    else:
        tk.messagebox.showinfo(
            "E-04", "Wrong username. Please enter the correct username and try again.")
        sys.exit()

def home():
    global title_frame,var,data,user_name, title_label, request_frame,var1,var2,var3,var4,var5, access_frame, Label_username, UseNameEntry, Label_iponame, IpoNameEntry, Label_lotsize, lotsizeEntry, label_pancolumn, PanEntry, label_startrow, StartRowEntry, label_endrow, StopRowEntry, label_websites, websites_menu, submit_button
    
    var = tk.IntVar()
    var1 = tk.IntVar(value=2)
    var2= tk.IntVar(value=3)
    var3 = tk.IntVar(value=2)
    var4 = tk.IntVar(value=2)
    var5 = tk.StringVar(value="chrome")
    
    title_frame = ctk.CTkFrame(root, width = 1020, height = 40 )
    title_frame.place(relx = 0.5, rely = 0, anchor = "n")

    title_label = ctk.CTkLabel(title_frame, text="IPO Allotment Checker", font=("Arial Black", 30))
    title_label.place(relx = 0.2, rely = 0, anchor = "n")
    
    ed_frame = ctk.CTkFrame(title_frame, width = 250, height = 30, fg_color='transparent')
    ed_frame.place(relx = 0.85, rely = 0.075, anchor = "n")
    if data == None:
        data = 'Inactive'
    ed_label = ctk.CTkLabel(ed_frame, text=f"License :{data}", font=("Arial Black", 15),fg_color='transparent')
    ed_label.place(relx = 0.5, rely = 0.12, anchor = "n")
    
    theme_frame = ctk.CTkFrame(title_frame, width = 20, height = 10, fg_color='transparent')
    theme_frame.place(relx = 0.975, rely = 0, anchor = "n")
    theme_icon = Image.open("brightness.png") 
    theme_icon = theme_icon.resize((20, 20), Image.BILINEAR)
    theme_icon_tk = ImageTk.PhotoImage(theme_icon)
    toggle_button = ctk.CTkButton(theme_frame, text ='', width=20,height = 20,font=("Arial Black", 18), image=theme_icon_tk,command=toggle_theme,border_width=1)
    toggle_button.image = theme_icon_tk
    toggle_button.pack(side='left', padx=6, pady=10)

    request_frame = ctk.CTkScrollableFrame(root, width=920, height=530, border_width=0,corner_radius=0)
    request_frame.place(relx = 1, rely = 0.085, anchor = "ne")

    access_frame = ctk.CTkFrame(root, width=80, height=530, border_width=0 ,corner_radius=0)
    access_frame.place(relx = 0, rely = 0.085, anchor = "nw")

    username_frame = ctk.CTkFrame(request_frame)
    username_frame.pack(side="top", pady=(10, 10))
    Label_username = ctk.CTkLabel(username_frame, text="User Name :", font=("Arial Black", 18))
    Label_username.pack(side="left", padx=(0, 380))
    UseNameEntry = ctk.CTkEntry(username_frame, width=350,font=("Century Gothic bold", 18),  corner_radius=0, border_width=0)
    UseNameEntry.pack(side="left",padx=(20, 0),pady=10)
    if user_name != None:
        # user_name = 'Inactive'
        UseNameEntry.insert(0, user_name)

    Iponame_frame = ctk.CTkFrame(request_frame)
    Iponame_frame.pack(side="top", pady=(10, 10))
    Label_iponame = ctk.CTkLabel(Iponame_frame, text="IPO Name :", font=("Arial Black", 18))
    Label_iponame.pack(side="left", padx=(0, 390))
    IpoNameEntry = ctk.CTkEntry(Iponame_frame, width=350,font=("Century Gothic bold", 18),  corner_radius=0, border_width=0)
    IpoNameEntry.pack(side="left",padx=(20, 0),pady=10)

    pancolumn_frame = ctk.CTkFrame(request_frame)
    pancolumn_frame.pack(side="top", pady=(10, 10))
    label_pancolumn = ctk.CTkLabel(pancolumn_frame, text="Pan No is in which column ?", font=("Arial Black", 18))
    label_pancolumn.pack(side="left", padx=(0, 245))
    PanEntry = ctk.CTkEntry(pancolumn_frame, width=350,font=("Century Gothic bold", 18), corner_radius=0, border_width=0)    
    PanEntry.pack(side="left",pady=10)
    PanEntry.insert(0, '1')
    
    start_frame = ctk.CTkFrame(request_frame)
    start_frame.pack(side="top", pady=(10, 10))
    label_startrow = ctk.CTkLabel(start_frame, text="From which Row to you want to start ?", font=("Arial Black", 18))
    label_startrow.pack(side="left", padx=(0, 140))
    StartRowEntry = ctk.CTkEntry(start_frame, width=350,font=("Century Gothic bold", 18),  corner_radius=0, border_width=0)
    StartRowEntry.pack(side="left",pady=10)
    StartRowEntry.insert(0, '2')

    end_frame = ctk.CTkFrame(request_frame)
    end_frame.pack(side="top", pady=(10, 10))
    label_endrow = ctk.CTkLabel(end_frame, text="From which Row to you want to stop ?", font=("Arial Black", 18))
    label_endrow.pack(side="left", padx=(0, 140))
    StopRowEntry = ctk.CTkEntry(end_frame, width=350, bg_color="#0f0f0f", font=("Century Gothic bold", 18),corner_radius=0, border_width=0)
    StopRowEntry.pack(side="left",pady=10)
    StopRowEntry.insert(0, 'MAX ROW')

    site_frame = ctk.CTkFrame(request_frame)
    site_frame.pack(side="top", pady=(10, 10))
    label_websites = ctk.CTkLabel(site_frame, text="Select the Website", font=("Arial Black", 18))
    label_websites.pack(side="left", padx=(0, 310))
    websites = ["Link Intime", "Kfintech", "SkyLine", "Bigshare", "Purva","Integrated","Maashitla","Cambridge"]
    def update_variable(selection):
        if selection == "Link Intime":
            var.set(2)
        elif selection == "Kfintech":
            var.set(1)
        elif selection == "SkyLine":
            var.set(4)
        elif selection == "Bigshare":
            var.set(3)
        elif selection == "Purva":
            var.set(5)
        elif selection == "Maashitla":
            var.set(6)
        elif selection == "Integrated":
            var.set(7)        
        elif selection == "Cambridge":
            var.set(8)        
    
    update_variable("Link Intime")
    websites_menu = ctk.CTkOptionMenu(site_frame, values=websites,font=("Century Gothic bold", 15), width=350,command=update_variable)
    websites_menu.pack(side="left",padx=(20, 0),pady=10)

    submit_button = ctk.CTkButton(request_frame, text="Submit", font=("Arial Black", 18),  border_width=1,command=submit)
    submit_button.pack(pady=20)

    home_icon = Image.open("home.png") 

    home_icon = home_icon.resize((30, 30), Image.BILINEAR)
    home_icon_tk = ImageTk.PhotoImage(home_icon)

    access_home  = ctk.CTkButton(access_frame, width = 280,text="Home",height=50, font=("Arial Black", 17), border_width=0, command=home,image=home_icon_tk,compound="top")
    access_home.image = home_icon_tk
    access_home.place(relx = 0.5, rely = 0.01, anchor = "n")

    Setting = Image.open("cogwheel.png") 

    Setting = Setting.resize((30, 30), Image.BILINEAR)
    Setting_tk = ImageTk.PhotoImage(Setting)

    access_advanced  = ctk.CTkButton(access_frame,width = 280, text="Setting", font=("Arial Black", 17), border_width=0, command=advanced,image=Setting_tk,compound="top")
    access_advanced.image = Setting_tk
    access_advanced.place(relx = 0.5, rely = 0.15, anchor = "n")
    
    Counter = Image.open("counter.png") 

    Counter = Counter.resize((30, 30), Image.BILINEAR)
    Counter_tk = ImageTk.PhotoImage(Counter)
    
    access_counter  = ctk.CTkButton(access_frame,width = 280, text="Counter", font=("Arial Black", 17), border_width=0, command=counter,image=Counter_tk,compound="top")
    access_counter.image = Counter_tk
    access_counter.place(relx = 0.5, rely = 0.29, anchor = "n")

def advanced():
    global var4,var3
    var4 = ctk.IntVar()
    var3 = ctk.IntVar()
    
    try:
        for widget in request_frame.winfo_children():
            widget.destroy()
    except:
        pass
    
    # mydb = dbconnection() 
        
    macid = WMI().Win32_ComputerSystemProduct()[0].UUID
    print(macid)  # B2523B8E-6805-582D-3D7A-796EE2191689
    GetMacid = f"select MacId from Customer where MacId = '{macid}'"
    df = pd.read_sql(GetMacid, con=mydb)

    try:
        result = df.iloc[0, 0]
    except:
        tk.messagebox.showinfo(
            "E-06", "This PC is not registered with your User Name. Please contact administration at arhamtechnologyindia@gmail.com to resolve this issue.\nWhatsApp: 7228882088")
        sys.exit()
    
    if macid == result:
        GetBigshareURL = f"select BigshareURL from Customer where MacId = '{macid}'"
        df2 = pd.read_sql(GetBigshareURL, con=mydb)
        Bigshareurl = df2.iloc[0, 0]
        try:
            var1.set(int(Bigshareurl))
        except:
            var1.set(1)
        
        GetKaryURL = f"select KaryURL from Customer where MacId = '{macid}'"
        df2 = pd.read_sql(GetKaryURL, con=mydb)
        karyurl = df2.iloc[0, 0]
        try:
            var2.set(int(karyurl))
        except:
            var2.set(1)
        
        
        #Get Retry Count
        GetRetryCount = f"select RetryCount from Customer where MacId = '{macid}'"
        df1 = pd.read_sql(GetRetryCount, con=mydb)
        RetryCount = int(df1.iloc[0, 0])
        try:
            var4.set(int(RetryCount))
        except:
            var4.set(1)
        
        #Multiple Pan Value
        GetMultiplePan = f"select MultiplePan from Customer where MacId = '{macid}'"
        dff = pd.read_sql(GetMultiplePan, con=mydb)
        MultiplePan = dff.iloc[0, 0]
        try:
            var3.set(int(MultiplePan))
        except:
            var3.set(1)
        
        GetBrowser = f"select Browser from Customer where MacId = '{macid}'"
        dfff = pd.read_sql(GetBrowser, con=mydb)
        Browser = dfff.iloc[0, 0]
        try:
            var5.set(Browser)
        except:
            var5.set("chrome")

    else:
        tk.messagebox.showinfo(
            "E-06", "This PC is not registered with your User Name. Please contact administration at arhamtechnologyindia@gmail.com to resolve this issue.\nWhatsApp: 7228882088")
        sys.exit()

    def apply():    
        mycursor = mydb.cursor()

        UpdateBigsharelink = f"Update Customer set BigshareURL = {var1.get()},KaryURL = {var2.get()},MultiplePan = {var3.get()},RetryCount = {var4.get()},Browser = '{var5.get()}' Where MacId = '{macid}'"
        mycursor.execute(UpdateBigsharelink)
        mydb.commit()
        home()
        
    kfintech_frame = ctk.CTkFrame(request_frame)
    kfintech_frame.pack(side="top", pady=(10, 10))
    label_websites = ctk.CTkLabel(kfintech_frame, text="Select the Kfintech Server", font=("Arial Black", 18))
    label_websites.pack(side="left", padx=(0, 180))
    websites = ["https://kosmic.kfintech.com/ipostatus/", "https://kcas.kfintech.com/ipostatus/", "https://kprism.kfintech.com/ipostatus/", "https://evault.kfintech.com/ipostatus/","https://rti.kfintech.com/ipostatus/"]
    def update_klink(selection):
        if selection == "https://kosmic.kfintech.com/ipostatus/":
            var2.set(1)
        elif selection == "https://kcas.kfintech.com/ipostatus/":
            var2.set(2)
        elif selection == "https://kprism.kfintech.com/ipostatus/":
            var2.set(3)
        elif selection == "https://evault.kfintech.com/ipostatus/":
            var2.set(4)
        elif selection == "https://rti.kfintech.com/ipostatus/":
            var2.set(5)
    
    v2 = (var2.get())-1
    v2 = websites[v2]
    websites_menu = ctk.CTkOptionMenu(kfintech_frame,values=websites,font=("Century Gothic bold", 15), width=350,command=update_klink)
    websites_menu.pack(side="left",padx=(20, 0),pady=10)
    websites_menu.set(v2)
    
    bigshare_frame = ctk.CTkFrame(request_frame)
    bigshare_frame.pack(side="top", pady=(10, 10))
    label_websites = ctk.CTkLabel(bigshare_frame, text="Select the Bigshare Server", font=("Arial Black", 18))
    label_websites.pack(side="left", padx=(0, 178))
    websites = ["https://ipo.bigshareonline.com/IPO_Status.html", "https://ipo1.bigshareonline.com/IPO_Status.html", "https://ipo2.bigshareonline.com/IPO_Status.html"]
    def update_blink(selection):
        if selection == "https://ipo.bigshareonline.com/IPO_Status.html":
            var1.set(1)
        elif selection == "https://ipo1.bigshareonline.com/IPO_Status.html":
            var1.set(2)
        elif selection == "https://ipo2.bigshareonline.com/IPO_Status.html":
            var1.set(3)
    
    v1 = (var1.get())-1
    v1 = websites[v1]
    websites_menu = ctk.CTkOptionMenu(bigshare_frame, values=websites,font=("Century Gothic bold", 13), width=350,command=update_blink)
    websites_menu.pack(side="left",padx=(20, 0),pady=10)
    websites_menu.set(v1)

    malti_frame = ctk.CTkFrame(request_frame)
    malti_frame.pack(side="top", pady=(10, 10))
    label_multiplepan = ctk.CTkLabel(malti_frame, text="Do you want to Check Multiple PAN ?", font=("Arial Black", 20) )
    label_multiplepan.pack(side="left",padx=(0, 165))
    radio_multiplepan1 = ctk.CTkRadioButton(malti_frame, text="Yes", variable=var3, value=1, font=("Arial Black", 20))
    radio_multiplepan1.pack(side="left",padx=(10, 0),pady=10)
    radio_multiplepan2 = ctk.CTkRadioButton(malti_frame, text="No", variable=var3, value=2, font=("Arial Black", 20))
    radio_multiplepan2.pack(side="left",padx=(30, 0),pady=10)

    
    retry_frame = ctk.CTkFrame(request_frame)
    retry_frame.pack(side="top", pady=(10, 10))
    label_retry= ctk.CTkLabel(retry_frame, text="How many retry do want to make", font=("Arial Black", 20) )
    label_retry.pack(side="left",padx=(0, 115))
    radio_retry1 = ctk.CTkRadioButton(retry_frame, text="0", variable=var4, value=1, font=("Arial Black", 20) )
    radio_retry1.pack(side="left",padx=(10, 0),pady=10)
    radio_retry2 = ctk.CTkRadioButton(retry_frame, text="1", variable=var4, value=2, font=("Arial Black", 20) )
    radio_retry2.pack(side="left",padx=(10, 0),pady=10)
    radio_retry3 = ctk.CTkRadioButton(retry_frame, text="2", variable=var4, value=3, font=("Arial Black", 20) )
    radio_retry3.pack(side="left",padx=(10, 0),pady=10)

    browsers_frame = ctk.CTkFrame(request_frame)
    browsers_frame.pack(side="top", pady=(10, 10))
    label_websites = ctk.CTkLabel(browsers_frame, text=" Select the Browser", font=("Arial Black", 18))
    label_websites.pack(side="left", padx=(0, 245))
    websites = ["chrome", "firefox", "edge"]
    def Browser_update(selection):
        if selection == "chrome":
            var5.set('chrome')
        elif selection == "firefox":
            var5.set('firefox')
        elif selection == "edge":
            var5.set('edge')
    
    Browser_update(var5)
    websites_menu = ctk.CTkOptionMenu(browsers_frame,variable=var5, values=websites,font=("Century Gothic bold", 18), width=350,command=Browser_update)
    websites_menu.pack(side="left",padx=(30, 0),pady=10)

    Apply_button = ctk.CTkButton(request_frame, text="Apply", font=("Arial Black", 20), border_width=1 ,command =apply)
    Apply_button.pack(pady=50)
    
def counter():
    global variable,variable1,variable3,variable2,variable4,variable5
    try:
        for widget in request_frame.winfo_children():
            widget.destroy()
    except:
        pass
    
    variable = ctk.StringVar()
    variable1 = ctk.StringVar()
    variable2 = ctk.StringVar()
    variable3 = ctk.StringVar()
    variable4 = ctk.StringVar()
    variable5 = ctk.StringVar()
    
    counter_frame = ctk.CTkFrame(request_frame)
    counter_frame.pack(side="top", pady=(10, 10),padx=(0,400))
    CounterLable = ctk.CTkLabel(counter_frame,font=("Arial Black", 18), text="COUNTER:")
    CounterLable.pack(side="left", padx=(0, 250),pady=10)
    CounterVariable = ctk.CTkLabel(counter_frame,font=("Century Gothic bold", 18), textvariable=variable)
    CounterVariable.pack(side="left", padx=(0, 0),pady=10)

    Alloted_frame = ctk.CTkFrame(request_frame)
    Alloted_frame.pack(side="top", pady=(10, 10),padx=(0,400))
    AllotedLable = ctk.CTkLabel(Alloted_frame,font=("Arial Black", 18), text="ALLOTTED:")
    AllotedLable.pack(side="left", padx=(0, 240),pady=10)
    AllotedVariabe = ctk.CTkLabel(Alloted_frame,font=("Century Gothic bold", 18), textvariable=variable1)
    AllotedVariabe.pack(side="left", padx=(0, 0),pady=10)

    notalloted_frame = ctk.CTkFrame(request_frame)
    notalloted_frame.pack(side="top", pady=(10, 10),padx=(0,400))
    NotAllotedLable = ctk.CTkLabel(notalloted_frame,font=("Arial Black", 18), text="NOT ALLOTTED:")
    NotAllotedLable.pack(side="left", padx=(0, 190),pady=10)
    NotAllotedVariabe = ctk.CTkLabel(notalloted_frame,font=("Century Gothic bold", 18), textvariable=variable2)
    NotAllotedVariabe.pack(side="left", padx=(0, 0),pady=10)

    Error_frame = ctk.CTkFrame(request_frame)
    Error_frame.pack(side="top", pady=(10, 10),padx=(0,400))
    ErrorLable = ctk.CTkLabel(Error_frame,font=("Arial Black", 18), text="ERROR:")
    ErrorLable.pack(side="left", padx=(0, 280),pady=10)
    ErrorVariable = ctk.CTkLabel(Error_frame, textvariable=variable3,font=("Century Gothic bold", 18),)
    ErrorVariable.pack(side="left", padx=(0, 0),pady=10)

    av_frame = ctk.CTkFrame(request_frame)
    av_frame.pack(side="top", pady=(10, 10),padx=(0,380))
    av = ctk.CTkLabel(av_frame,font=("Arial Black", 18), text="AVERAGE SHARE PER APP.:")
    av.pack(side="left", padx=(0, 80),pady=10)
    av1 = ctk.CTkLabel(av_frame, textvariable=variable4,font=("Century Gothic bold", 18),)
    av1.pack(side="left",padx=(0, 0),pady=10)

    Total_frame = ctk.CTkFrame(request_frame)
    Total_frame.pack(side="top", pady=(10, 10),padx=(0,400))
    TotalSharesLabel = ctk.CTkLabel(Total_frame, font=("Arial Black", 18),text="TOTAL SHARES ALLOTTED:")
    TotalSharesLabel.pack(side="left", padx=(0, 80),pady=10)
    total = ctk.CTkLabel(Total_frame, textvariable=variable5,font=("Century Gothic bold", 18),)
    total.pack(side="left",padx=(0, 0),pady=10)

home()
root.resizable(False, False)
root.mainloop()